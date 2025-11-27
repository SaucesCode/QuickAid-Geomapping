# Standard library
import io
import os
import json
import base64
import hashlib
from io import BytesIO
from pathlib import Path
from PIL import Image as PILImage
from django.core.cache import cache
from datetime import date, datetime, timedelta


import seaborn as sns
from dateutil.relativedelta import relativedelta
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    Frame, Image, KeepTogether, PageBreak, PageTemplate, Paragraph,
    SimpleDocTemplate, Spacer, Table, TableStyle
)

# Django
from django.conf import settings
from django.db.models import (
    Avg, Count, DurationField, ExpressionWrapper, F,
    IntegerField, Max, Q
)
from django.db.models.functions import (
    ExtractHour, ExtractYear, TruncDate, TruncMonth
)
from django.utils import timezone
from openpyxl.formatting.rule import ColorScaleRule


# Data & Charts
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_pdf import PdfPages
import numpy as np

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, PageBreak, Image, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel Generation
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Local
from .models import Applicant, BackgroundInfo, StaffActivityLog
from .utils import apply_applicant_filters

def apply_common_filters(qs, filters):
    """
    Apply common filters to queryset
    Extracted from views.py to avoid circular imports
    """
    start_date = filters.get('start_date')
    end_date = filters.get('end_date')
    assistance_types = filters.get('assistance_types', [])
    cities = filters.get('cities', [])
    barangays = filters.get('barangays', [])
    
    if start_date and end_date:
        qs = qs.filter(date_filled__date__range=[start_date, end_date])
    if assistance_types:
        qs = qs.filter(type_of_assistance__in=assistance_types)
    if cities:
        qs = qs.filter(background_info__barangay__city__name__in=cities)
    if barangays:
        qs = qs.filter(background_info__barangay__name__in=barangays)
    
    return qs


class AnalyticsDataCollector:
    """Collects all analytics data based on filters"""
    
    def __init__(self, filters):
        self.filters = filters
        self.start_date = filters.get('start_date')
        self.end_date = filters.get('end_date')
        self.city = filters.get('city')
        self.cities = filters.get('cities', [])
        self.barangays = filters.get('barangays', [])
        self.assistance_types = filters.get('assistance_types', [])
        
    def _get_base_queryset(self):
        """Get base applicant queryset with filters applied"""
        qs = Applicant.objects.filter(is_archived=False)
        
        if self.start_date and self.end_date:
            qs = qs.filter(date_filled__date__range=[self.start_date, self.end_date])
        if self.assistance_types:
            qs = qs.filter(type_of_assistance__in=self.assistance_types)
        if self.city:
            qs = qs.filter(background_info__barangay__city__name=self.city)
        if self.cities:
            qs = qs.filter(background_info__barangay__city__name__in=self.cities)
        if self.barangays:
            qs = qs.filter(background_info__barangay__name__in=self.barangays)
            
        return qs
    
    def collect_summary_metrics(self):
        """Dashboard KPIs"""
        base_qs = self._get_base_queryset()
        total = base_qs.count()
        
        # Average processing time
        avg_time = base_qs.exclude(
            created_at__isnull=True, date_filled__isnull=True
        ).aggregate(
            avg_time=Avg(ExpressionWrapper(
                F('date_filled') - F('created_at'),
                output_field=DurationField()
            ))
        )['avg_time']
        avg_minutes = round(avg_time.total_seconds() / 60, 1) if avg_time else 0
        
        # Most common assistance type
        most_common = base_qs.values('type_of_assistance').annotate(
            count=Count('id')
        ).order_by('-count').first()
        
        # Top barangay
        top_barangay = base_qs.values('background_info__barangay__name').annotate(
            count=Count('id')
        ).order_by('-count').first()
        
        # Growth rate
        today = timezone.localdate()
        start_of_this_month = today.replace(day=1)
        start_of_last_month = (start_of_this_month - timedelta(days=1)).replace(day=1)
        end_of_last_month = start_of_this_month - timedelta(days=1)
        
        this_month = base_qs.filter(date_filled__date__gte=start_of_this_month).count()
        last_month = base_qs.filter(
            date_filled__date__range=[start_of_last_month, end_of_last_month]
        ).count()
        growth_rate = ((this_month - last_month) / last_month * 100) if last_month > 0 else 0
        
        return {
            'total_applicants': total,
            'avg_processing_minutes': avg_minutes,
            'most_common_type': most_common['type_of_assistance'] if most_common else 'N/A',
            'top_barangay': top_barangay['background_info__barangay__name'] if top_barangay else 'N/A',
            'growth_rate': round(growth_rate, 2),
            'this_month': this_month,
            'last_month': last_month
        }
    
    def collect_geographic_data(self):
        """Geographic insights"""
        base_qs = self._get_base_queryset()
        
        # Top 10 barangays
        top_barangays = list(
            base_qs.values('background_info__barangay__name')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        
        # By city
        by_city = list(
            base_qs.values('background_info__barangay__city__name')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        
        # Approval rates by location
        by_location = list(
            base_qs.values('background_info__barangay__city__name')
            .annotate(
                total=Count('id'),
                approved=Count('id', filter=Q(approvals__isnull=False))
            )
            .order_by('-total')
        )
        
        return {
            'top_barangays': top_barangays,
            'by_city': by_city,
            'approval_by_location': by_location
        }
    
    def collect_demographic_data(self):
        """Demographics analysis"""
        base_qs = self._get_base_queryset()
        today = date.today()
        
        # Gender
        by_gender = list(
            base_qs.values('background_info__sex')
            .annotate(count=Count('id'))
        )
        
        # Civil status
        by_civil_status = list(
            base_qs.values('background_info__civil_status')
            .annotate(count=Count('id'))
        )
        
        # Age groups
        qs_with_age = base_qs.annotate(
            age=ExpressionWrapper(
                today.year - ExtractYear('background_info__birthday'),
                output_field=IntegerField()
            )
        )
        age_groups = {
            '0-17': qs_with_age.filter(age__lte=17).count(),
            '18-25': qs_with_age.filter(age__gte=18, age__lte=25).count(),
            '26-35': qs_with_age.filter(age__gte=26, age__lte=35).count(),
            '36-45': qs_with_age.filter(age__gte=36, age__lte=45).count(),
            '46-60': qs_with_age.filter(age__gte=46, age__lte=60).count(),
            '60+': qs_with_age.filter(age__gt=60).count(),
        }
        
        # Occupation
        by_occupation = list(
            base_qs.values('background_info__occupation')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        
        return {
            'by_gender': by_gender,
            'by_civil_status': by_civil_status,
            'age_groups': age_groups,
            'by_occupation': by_occupation
        }
    
    def collect_economic_data(self):
        """Economic analysis"""
        base_qs = self._get_base_queryset().exclude(background_info__monthly_income=0)
        
        income_ranges = [
            (0, 10000, 'Below 10,000'),
            (10001, 20000, '10,001 - 20,000'),
            (20001, 30000, '20,001 - 30,000'),
            (30001, 40000, '30,001 - 40,000'),
            (40001, 50000, '40,001 - 50,000'),
            (50001, 100000, '50,001 - 100,000'),
            (100001, None, 'Above 100,000')
        ]
        
        distribution = []
        for min_income, max_income, label in income_ranges:
            if max_income is None:
                count = base_qs.filter(background_info__monthly_income__gte=min_income).count()
            else:
                count = base_qs.filter(
                    background_info__monthly_income__gte=min_income,
                    background_info__monthly_income__lt=max_income
                ).count()
            distribution.append({'range': label, 'count': count})
        
        return {'income_distribution': distribution}
    
    def collect_trends_data(self):
        """Trends over time"""
        base_qs = self._get_base_queryset()
        
        # Monthly trends (last 12 months)
        today = timezone.now().date()
        start_date = today.replace(day=1) - relativedelta(months=11)
        monthly = list(
            base_qs.filter(date_filled__gte=start_date)
            .annotate(month=TruncMonth('date_filled'))
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        )
        
        # Yearly trends
        yearly = list(
            base_qs.annotate(year=ExtractYear('date_filled'))
            .values('year')
            .annotate(count=Count('id'))
            .order_by('year')
        )
        
        # Assistance type distribution
        by_assistance = list(
            base_qs.values('type_of_assistance')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        
        # Assistance types over time (monthly)
        assistance_over_time = list(
            base_qs.annotate(month=TruncMonth('date_filled'))
            .values('month', 'type_of_assistance')
            .annotate(count=Count('id'))
            .order_by('month')
        )
        
        return {
            'monthly': monthly,
            'yearly': yearly,
            'by_assistance': by_assistance,
            'assistance_over_time': assistance_over_time
        }
    
    def collect_performance_data(self):
        """Performance & productivity metrics"""
        base_qs = self._get_base_queryset()
        
        # Staff productivity
        staff_productivity = list(
            base_qs.values('staff__username')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        
        # Processing time by assistance type
        processing_by_type = list(
            base_qs.exclude(created_at__isnull=True, date_filled__isnull=True)
            .annotate(
                duration=ExpressionWrapper(
                    F('date_filled') - F('created_at'),
                    output_field=DurationField()
                )
            )
            .values('type_of_assistance')
            .annotate(avg_time=Avg('duration'))
        )
        
        for item in processing_by_type:
            if item['avg_time']:
                item['avg_minutes'] = round(item['avg_time'].total_seconds() / 60, 1)
            else:
                item['avg_minutes'] = 0  # ensure the key exists
            
        # Activity heatmap (by hour)
        activity_heatmap = list(
            base_qs.annotate(hour=ExtractHour('date_filled'))
            .values('hour')
            .annotate(count=Count('id'))
            .order_by('hour')
        )
        
        return {
            'staff_productivity': staff_productivity,
            'processing_by_type': processing_by_type,
            'activity_heatmap': activity_heatmap
        }
    
    def collect_all(self):
        """Collect all analytics data"""
        return {
            'summary': self.collect_summary_metrics(),
            'geographic': self.collect_geographic_data(),
            'demographics': self.collect_demographic_data(),
            'economics': self.collect_economic_data(),
            'trends': self.collect_trends_data(),
            'performance': self.collect_performance_data()
        }


class InsightsGenerator:
    """Generates narrative insights from analytics data"""
    
    def __init__(self, data):
        self.data = data
    
    def generate_executive_summary(self):
        """Generate high-level summary"""
        summary = self.data['summary']
        
        insights = []
        insights.append(
            f"During the reporting period, a total of {summary['total_applicants']:,} "
            f"applicants were processed with an average processing time of "
            f"{summary['avg_processing_minutes']} minutes per application."
        )
        
        if summary['growth_rate'] > 0:
            insights.append(
                f"Applications showed a {summary['growth_rate']}% increase compared to "
                f"the previous month ({summary['this_month']} vs {summary['last_month']})."
            )
        elif summary['growth_rate'] < 0:
            insights.append(
                f"Applications decreased by {abs(summary['growth_rate'])}% compared to "
                f"the previous month ({summary['this_month']} vs {summary['last_month']})."
            )
        else:
            insights.append("Applications remained stable compared to the previous month.")
        
        insights.append(
            f"The most requested assistance type was {summary['most_common_type']}, "
            f"with {summary['top_barangay']} recording the highest number of applications."
        )
        
        return " ".join(insights)
    
    def generate_geographic_insights(self):
        """Generate geographic insights"""
        geo = self.data['geographic']
        insights = []
        
        if geo['top_barangays']:
            top = geo['top_barangays'][0]
            insights.append(
                f"• {top['background_info__barangay__name']} leads with {top['count']} applications"
            )
        
        if len(geo['by_city']) > 1:
            cities_sorted = sorted(geo['by_city'], key=lambda x: x['count'], reverse=True)
            top_city = cities_sorted[0]
            insights.append(
                f"• {top_city['background_info__barangay__city__name']} has the highest city-wide "
                f"applications with {top_city['count']} total"
            )
        
        # Approval rate analysis
        if geo['approval_by_location']:
            for loc in geo['approval_by_location'][:3]:
                if loc['total'] > 0:
                    rate = (loc['approved'] / loc['total']) * 100
                    insights.append(
                        f"• {loc['background_info__barangay__city__name']}: "
                        f"{rate:.1f}% approval rate ({loc['approved']}/{loc['total']})"
                    )
        
        return insights
    
    def generate_demographic_insights(self):
        """Generate demographic insights"""
        demo = self.data['demographics']
        insights = []
        
        # Gender distribution
        if demo['by_gender']:
            gender_sorted = sorted(demo['by_gender'], key=lambda x: x['count'], reverse=True)
            total = sum(g['count'] for g in gender_sorted)
            if total > 0:
                top_gender = gender_sorted[0]
                pct = (top_gender['count'] / total) * 100
                insights.append(
                    f"• {top_gender['background_info__sex']} applicants comprise {pct:.1f}% "
                    f"of total applications"
                )
        
        # Age groups
        age_groups = demo['age_groups']
        if age_groups:
            dominant_age = max(age_groups.items(), key=lambda x: x[1])
            insights.append(
                f"• Age group {dominant_age[0]} is most represented with {dominant_age[1]} applicants"
            )
        
        # Civil status
        if demo['by_civil_status']:
            civil_sorted = sorted(demo['by_civil_status'], key=lambda x: x['count'], reverse=True)
            if civil_sorted:
                insights.append(
                    f"• {civil_sorted[0]['background_info__civil_status']} applicants: "
                    f"{civil_sorted[0]['count']}"
                )
        
        return insights
    
    def generate_trend_insights(self):
        """Generate trend insights"""
        trends = self.data['trends']
        insights = []
        
        # Monthly trend analysis
        if len(trends['monthly']) >= 2:
            latest = trends['monthly'][-1]['count']
            previous = trends['monthly'][-2]['count']
            change = ((latest - previous) / previous * 100) if previous > 0 else 0
            
            if change > 10:
                insights.append(f"• Strong growth: Applications increased {change:.1f}% month-over-month")
            elif change < -10:
                insights.append(f"• Declining trend: Applications decreased {abs(change):.1f}% month-over-month")
            else:
                insights.append("• Stable trend: Applications remain consistent month-over-month")
        
        # Assistance type trends
        if trends['by_assistance']:
            top_3 = trends['by_assistance'][:3]
            insights.append("• Top assistance types: " + ", ".join(
                f"{t['type_of_assistance']} ({t['count']})" for t in top_3
            ))
        
        return insights
    
    def generate_performance_insights(self):
        """Generate performance insights"""
        perf = self.data['performance']
        insights = []
        
        # Staff productivity
        if perf['staff_productivity']:
            top_staff = perf['staff_productivity'][0]
            insights.append(
                f"• Most productive staff: {top_staff['staff__username']} "
                f"({top_staff['count']} applications processed)"
            )
        
        # Processing time analysis
        if perf['processing_by_type']:
            fastest = min(perf['processing_by_type'], key=lambda x: x.get('avg_minutes', float('inf')))
            slowest = max(perf['processing_by_type'], key=lambda x: x.get('avg_minutes', 0))
            
            insights.append(
                f"• Fastest processing: {fastest['type_of_assistance']} "
                f"({fastest.get('avg_minutes', 0):.1f} min avg)"
            )
            insights.append(
                f"• Slowest processing: {slowest['type_of_assistance']} "
                f"({slowest.get('avg_minutes', 0):.1f} min avg)"
            )
        
        # Peak hours
        if perf['activity_heatmap']:
            peak_hour = max(perf['activity_heatmap'], key=lambda x: x['count'])
            insights.append(
                f"• Peak application hour: {peak_hour['hour']}:00 ({peak_hour['count']} applications)"
            )
        
        return insights
    
    def generate_recommendations(self):
        """Generate actionable recommendations"""
        recommendations = []
        summary = self.data['summary']
        
        # Processing time recommendations
        if summary['avg_processing_minutes'] > 10:
            recommendations.append(
                "• Consider streamlining the application process to reduce processing time"
            )
        
        # Geographic recommendations
        geo = self.data['geographic']
        if geo['top_barangays'] and len(geo['top_barangays']) > 0:
            top = geo['top_barangays'][0]
            recommendations.append(
                f"• Allocate additional resources to {top['background_info__barangay__name']} "
                f"due to high application volume"
            )
        
        # Growth recommendations
        if summary['growth_rate'] > 20:
            recommendations.append(
                "• Rapid growth detected - ensure adequate staffing and resources"
            )
        
        return recommendations
    
    def generate_all(self):
        """Generate all insights"""
        return {
            'executive_summary': self.generate_executive_summary(),
            'geographic': self.generate_geographic_insights(),
            'demographic': self.generate_demographic_insights(),
            'trends': self.generate_trend_insights(),
            'performance': self.generate_performance_insights(),
            'recommendations': self.generate_recommendations()
        }


class ChartGenerator:
    """Generates professional, publication-ready charts"""
    
    def __init__(self):
        # Set professional style
        sns.set_style("whitegrid")
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica']
        
        # DSWD Government colors
        self.primary_color = '#0066cc'
        self.secondary_color = '#4a90e2'
        self.accent_color = '#00cc66'
        
        # Professional color palette
        self.colors = [
            '#0066cc', '#00cc66', '#ff9800', '#9c27b0', 
            '#f44336', '#00bcd4', '#4caf50', '#ff5722'
        ]
    
    def _apply_professional_style(self, ax, title, xlabel, ylabel):
        """Apply consistent professional styling"""
        ax.set_title(title, fontsize=16, fontweight='bold', pad=20, color='#2c3e50')
        ax.set_xlabel(xlabel, fontsize=12, fontweight='600', color='#34495e')
        ax.set_ylabel(ylabel, fontsize=12, fontweight='600', color='#34495e')
        ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#bdc3c7')
        ax.spines['bottom'].set_color('#bdc3c7')
        ax.tick_params(colors='#7f8c8d', labelsize=10)
    
    def create_bar_chart(self, data, x_field, y_field, title, xlabel, ylabel, top_n=None):

        # Extract data
        x_values = [str(item.get(x_field, "")) for item in data]
        y_values = [int(item.get(y_field, 0) or 0) for item in data]

        # Sort and limit data if top_n is provided
        if top_n:
            combined = sorted(zip(x_values, y_values), key=lambda x: x[1], reverse=True)[:top_n]
            x_values, y_values = zip(*combined) if combined else ([], [])

        # Prevent empty or zero-only data from crashing
        if not y_values or sum(y_values) == 0:
            fig, ax = plt.subplots(figsize=(6, 3.5))
            ax.text(0.5, 0.5, "No data available", ha='center', va='center', fontsize=12, color='gray')
            ax.axis("off")
            return fig

        # Safe color scaling
        max_val = max(y_values)
        if max_val == 0:
            max_val = 1

        bar_colors = [plt.cm.Blues(0.4 + 0.5 * (val / max_val)) for val in y_values]

        # Create chart
        fig, ax = plt.subplots(figsize=(6, 3.5))
        ax.barh(x_values, y_values, color=bar_colors)
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.tick_params(axis='y', labelsize=9)
        plt.tight_layout()
        return fig

    
    def create_pie_chart(self, data, label_field, value_field, title):
        """Enhanced pie chart"""
        fig, ax = plt.subplots(figsize=(11, 9))
        
        labels = [str(item[label_field]) for item in data]
        values = [item[value_field] for item in data]
        
        filtered = [(l, v) for l, v in zip(labels, values) if v > 0]
        if not filtered:
            plt.close(fig)
            return None
        
        labels, values = zip(*filtered)
        colors_to_use = self.colors[:len(values)]
        
        wedges, texts, autotexts = ax.pie(
            values, labels=labels, 
            autopct=lambda pct: f'{pct:.1f}%\n({int(pct/100.*sum(values)):,})',
            colors=colors_to_use,
            startangle=90,
            explode=[0.05 if i == 0 else 0 for i in range(len(values))],
            shadow=True,
            textprops={'fontsize': 10, 'fontweight': '600'}
        )
        
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontsize(9)
            autotext.set_fontweight('bold')
        
        ax.set_title(title, fontsize=16, fontweight='bold', pad=20, color='#2c3e50')
        plt.tight_layout()
        
        return fig
    
    def create_line_chart(self, data, x_field, y_field, title, xlabel, ylabel):
        """Enhanced line chart with area fill"""
        fig, ax = plt.subplots(figsize=(14, 7))
        
        sorted_data = sorted(data, key=lambda x: x[x_field])
        x_values = [item[x_field] for item in sorted_data]
        y_values = [item[y_field] for item in sorted_data]
        
        # Main line
        ax.plot(x_values, y_values, marker='o', color=self.primary_color, 
                linewidth=3, markersize=8, markerfacecolor='white', 
                markeredgewidth=2, markeredgecolor=self.primary_color)
        
        # Area fill
        ax.fill_between(range(len(x_values)), y_values, alpha=0.2, color=self.primary_color)
        
        # Value labels
        for i, (x, y) in enumerate(zip(range(len(x_values)), y_values)):
            ax.text(x, y + max(y_values)*0.03, f'{int(y):,}', 
                   ha='center', va='bottom', fontsize=9, fontweight='bold', color='#2c3e50')
        
        # Format x-axis
        if isinstance(x_values[0], datetime):
            x_labels = [x.strftime('%b %Y') for x in x_values]
        else:
            x_labels = [str(x) for x in x_values]
        
        ax.set_xticks(range(len(x_values)))
        ax.set_xticklabels(x_labels, rotation=45, ha='right')
        
        self._apply_professional_style(ax, title, xlabel, ylabel)
        plt.tight_layout()
        
        return fig
    
    # Keep all other methods (create_stacked_bar_chart, create_heatmap, generate_all_charts)
    # exactly as they are in your original file!
    
    def create_stacked_bar_chart(self, data, title):
        """Stacked bar chart for assistance types over time"""
        fig, ax = plt.subplots(figsize=(14, 7))
        
        # Organize data by month and type
        months = sorted(set(item['month'] for item in data))
        types = sorted(set(item['type_of_assistance'] for item in data))
        
        # Create data matrix
        data_matrix = {t: [] for t in types}
        for month in months:
            for t in types:
                count = next((item['count'] for item in data 
                            if item['month'] == month and item['type_of_assistance'] == t), 0)
                data_matrix[t].append(count)
        
        # Plot stacked bars
        bottom = np.zeros(len(months))
        for idx, t in enumerate(types):
            ax.bar([m.strftime('%b %Y') if hasattr(m, 'strftime') else str(m) for m in months],
                  data_matrix[t], bottom=bottom, label=t, 
                  color=self.colors[idx % len(self.colors)])
            bottom += np.array(data_matrix[t])
        
        ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
        ax.set_xlabel('Month', fontsize=11)
        ax.set_ylabel('Number of Applications', fontsize=11)
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        
        return fig
    
    def create_heatmap(self, data, title):
        """Heatmap for hourly activity"""
        fig, ax = plt.subplots(figsize=(12, 6))
        
        hours = list(range(24))
        counts = [0] * 24
        
        for item in data:
            counts[item['hour']] = item['count']
        
        # Create heatmap-style bar chart
        colors_map = plt.cm.Blues(np.linspace(0.3, 1, 24))
        bars = ax.bar(hours, counts, color=colors_map, edgecolor='black', linewidth=0.5)
        
        # Add value labels
        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}', ha='center', va='bottom', fontsize=8)
        
        ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
        ax.set_xlabel('Hour of Day', fontsize=11)
        ax.set_ylabel('Number of Applications', fontsize=11)
        ax.set_xticks(hours)
        ax.set_xticklabels([f'{h:02d}:00' for h in hours], rotation=45, ha='right')
        plt.tight_layout()
        
        return fig
    
    def generate_all_charts(self, data):
        """Generate all charts and return as dict of figures"""
        charts = {}
        
        # Geographic charts
        if data['geographic']['top_barangays']:
            charts['top_barangays'] = self.create_bar_chart(
                data['geographic']['top_barangays'],
                'background_info__barangay__name', 'count',
                'Top 10 Barangays by Applications',
                'Barangay', 'Number of Applications'
            )
        
        if data['geographic']['by_city']:
            charts['by_city'] = self.create_pie_chart(
                data['geographic']['by_city'],
                'background_info__barangay__city__name', 'count',
                'Applications by City'
            )
        
        # Demographic charts
        if data['demographics']['by_gender']:
            charts['by_gender'] = self.create_pie_chart(
                data['demographics']['by_gender'],
                'background_info__sex', 'count',
                'Applications by Gender'
            )
        
        if data['demographics']['age_groups']:
            age_data = [{'group': k, 'count': v} for k, v in data['demographics']['age_groups'].items()]
            charts['age_groups'] = self.create_bar_chart(
                age_data, 'group', 'count',
                'Applications by Age Group',
                'Age Group', 'Number of Applications'
            )
        
        if data['demographics']['by_civil_status']:
            charts['civil_status'] = self.create_bar_chart(
                data['demographics']['by_civil_status'],
                'background_info__civil_status', 'count',
                'Applications by Civil Status',
                'Civil Status', 'Number of Applications'
            )
        
        # Economic charts
        if data['economics']['income_distribution']:
            charts['income'] = self.create_bar_chart(
                data['economics']['income_distribution'],
                'range', 'count',
                'Income Distribution',
                'Income Range (PHP)', 'Number of Applicants'
            )
        
        # Trends charts
        if data['trends']['monthly']:
            charts['monthly_trend'] = self.create_line_chart(
                data['trends']['monthly'],
                'month', 'count',
                'Monthly Application Trends (Last 12 Months)',
                'Month', 'Number of Applications'
            )
        
        if data['trends']['yearly']:
            charts['yearly_trend'] = self.create_bar_chart(
                data['trends']['yearly'],
                'year', 'count',
                'Yearly Application Trends',
                'Year', 'Number of Applications', top_n=20
            )
        
        if data['trends']['by_assistance']:
            charts['assistance_types'] = self.create_bar_chart(
                data['trends']['by_assistance'],
                'type_of_assistance', 'count',
                'Applications by Assistance Type',
                'Type of Assistance', 'Number of Applications'
            )
        
        if data['trends']['assistance_over_time']:
            charts['assistance_over_time'] = self.create_stacked_bar_chart(
                data['trends']['assistance_over_time'],
                'Assistance Types Over Time'
            )
        
        # Performance charts
        if data['performance']['staff_productivity']:
            charts['staff_productivity'] = self.create_bar_chart(
                data['performance']['staff_productivity'],
                'staff__username', 'count',
                'Staff Productivity (Top 10)',
                'Staff Member', 'Applications Processed'
            )
        
        if data['performance']['processing_by_type']:
            charts['processing_time'] = self.create_bar_chart(
                data['performance']['processing_by_type'],
                'type_of_assistance', 'avg_minutes',
                'Average Processing Time by Assistance Type',
                'Type of Assistance', 'Average Time (minutes)'
            )
        
        if data['performance']['activity_heatmap']:
            charts['activity_heatmap'] = self.create_heatmap(
                data['performance']['activity_heatmap'],
                'Application Activity by Hour of Day'
            )
        
        return charts


class PDFReportGenerator:
    """
    Professional Corporate PDF Report Generator
    Matches 6-sheet Excel structure with clean, modern design
    """
    
    def __init__(self, data, insights, charts, branding, filters):
        self.data = data
        self.insights = insights
        self.charts = charts
        self.branding = branding
        self.filters = filters
        
        # Document metadata
        self.doc_ref_number = branding.get('doc_ref_number', 
                                          f"ANALYTICS-{datetime.now().strftime('%Y%m%d-%H%M')}")
        self.office_name = branding.get('office_name', 'DSWD AICS Analytics')
        self.prepared_by = branding.get('prepared_by', 'Data Analytics Team')
        self.reviewed_by = branding.get('reviewed_by', 'Operations Manager')
        self.approved_by = branding.get('approved_by', 'Director')
        self.effectivity_date = branding.get('effectivity_date', 
                                            datetime.now().strftime('%B %d, %Y'))
        
        # Corporate colors (light theme matching Excel)
        self.primary_color = colors.HexColor('#0066cc')
        self.secondary_color = colors.HexColor('#003366')
        self.light_blue = colors.HexColor('#e3f2fd')
        self.accent_color = colors.HexColor('#00cc66')
        self.light_gray = colors.HexColor('#f5f5f5')
        
        self.styles = getSampleStyleSheet()
        self._setup_corporate_styles()
        
        # Page tracking
        self.page_count = 0

    def _format_location_filter(self):
        """Format location and assistance type filters for display"""
        city = self.filters.get('city')
        cities = self.filters.get('cities', [])
        barangays = self.filters.get('barangays', [])
        assistance_types = self.filters.get('assistance_types', [])
        
        result = []
        
        # City/Cities
        if city:
            result.append(f"City: {city}")
        elif cities:
            if len(cities) == 1:
                result.append(f"City: {cities[0]}")
            elif len(cities) <= 3:
                result.append(f"Cities: {', '.join(cities)}")
            else:
                result.append(f"Cities: {', '.join(cities[:3])} and {len(cities) - 3} more")
        else:
            result.append("Location: All Cities")
        
        # Barangays
        if barangays:
            if len(barangays) == 1:
                result.append(f"Barangay: {barangays[0]}")
            elif len(barangays) <= 3:
                result.append(f"Barangays: {', '.join(barangays)}")
            else:
                result.append(f"Barangays: {', '.join(barangays[:3])} and {len(barangays) - 3} more")
        else:
            if city or cities:
                result.append("Barangays: All Barangays")
        
        # Assistance Types
        if assistance_types:
            if len(assistance_types) == 1:
                result.append(f"Type: {assistance_types[0]}")
            elif len(assistance_types) <= 3:
                result.append(f"Types: {', '.join(assistance_types)}")
            else:
                result.append(f"Types: {', '.join(assistance_types[:2])} and {len(assistance_types) - 2} more")
        else:
            result.append("Assistance Types: All Types")
        
        return result
    
    def _setup_corporate_styles(self):
        """Setup clean corporate document styles"""
        
        # Helper function to safely add or update styles
        def add_or_update_style(name, **kwargs):
            if name in self.styles:
                # Update existing style
                style = self.styles[name]
                for key, value in kwargs.items():
                    setattr(style, key, value)
            else:
                # Add new style
                self.styles.add(ParagraphStyle(name=name, **kwargs))
        
        # Title Style
        add_or_update_style(
            'CorporateTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            textColor=self.primary_color,
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
            leading=24
        )
        
        # Section Header
        add_or_update_style(
            'SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=self.primary_color,
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold',
            borderWidth=1,
            borderColor=self.primary_color,
            borderPadding=6,
            backColor=self.light_blue,
            leading=18
        )
        
        # Subsection
        add_or_update_style(
            'SubSection',
            parent=self.styles['Heading3'],
            fontSize=12,
            textColor=self.secondary_color,
            spaceAfter=8,
            spaceBefore=12,
            fontName='Helvetica-Bold',
            leading=16
        )
        
        # Body Text (use unique name)
        add_or_update_style(
            'CorporateBody',  # Changed from 'BodyText'
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=8,
            alignment=TA_JUSTIFY,
            leading=14,
            fontName='Helvetica',
            textColor=colors.HexColor('#333333')
        )
        
        # Bullet Points
        add_or_update_style(
            'CorporateBullet',  # Changed from 'BulletPoint'
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            leftIndent=20,
            bulletIndent=10,
            leading=14,
            fontName='Helvetica'
        )
        
        # Small Text
        add_or_update_style(
            'CorporateSmall',  # Changed from 'SmallText'
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#666666'),
            alignment=TA_LEFT,
            leading=10,
            fontName='Helvetica'
        )
        
        # Signature Block
        add_or_update_style(
            'SignatureBlock',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=4,
            alignment=TA_CENTER,
            fontName='Helvetica'
        )
        
        # Table Cell Style
        add_or_update_style(
            'CorporateTableCell',  # Changed from 'TableCell'
            parent=self.styles['Normal'],
            fontSize=9,
            leading=12,
            fontName='Helvetica'
        )

    def _corporate_header_footer(self, canvas, doc):
        """Clean corporate header and footer on every page"""
        canvas.saveState()
        width, height = A4
        
        # ==================== HEADER ====================
        # Simple header bar
        canvas.setFillColor(self.light_blue)
        canvas.rect(0, height - 0.6*inch, width, 0.6*inch, fill=True, stroke=False)
        
        # Organization name (left)
        canvas.setFillColor(self.primary_color)
        canvas.setFont('Helvetica-Bold', 12)
        canvas.drawString(0.5*inch, height - 0.35*inch, 
                         self.branding.get('organization_name', 'DSWD Analytics Report'))
        
        # Document reference (right)
        canvas.setFont('Helvetica', 9)
        canvas.setFillColor(self.secondary_color)
        canvas.drawRightString(width - 0.5*inch, height - 0.35*inch,
                              f"Ref: {self.doc_ref_number}")
        
        # Classification (right, below ref)
        canvas.setFont('Helvetica-Bold', 8)
        canvas.drawRightString(width - 0.5*inch, height - 0.5*inch,
                              "OFFICIAL USE ONLY")
        
        # ==================== FOOTER ====================
        # Footer line
        canvas.setStrokeColor(self.light_blue)
        canvas.setLineWidth(2)
        canvas.line(0.5*inch, 0.6*inch, width - 0.5*inch, 0.6*inch)
        
        # Generation info (left)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.setFont('Helvetica', 8)
        canvas.drawString(0.5*inch, 0.4*inch,
                         f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
        
        # Page number (center)
        canvas.setFont('Helvetica', 9)
        canvas.drawCentredString(width/2, 0.4*inch,
                                f"Page {doc.page}")
        
        # Version info (right)
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(width - 0.5*inch, 0.4*inch,
                              f"Valid until: {self.effectivity_date}")
        
        canvas.restoreState()
    
    def _create_table(self, data, col_widths=None, has_header=True, zebra=True):
        """Create clean corporate table with light styling"""
        table = Table(data, colWidths=col_widths)
        
        style_list = [
            # Header row (light blue, not too bold)
            ('BACKGROUND', (0, 0), (-1, 0 if has_header else -1), self.light_blue),
            ('TEXTCOLOR', (0, 0), (-1, 0 if has_header else -1), self.secondary_color),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0 if has_header else -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0 if has_header else -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            
            # Light grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, self.primary_color),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]
        
        # Light zebra striping
        if zebra and len(data) > 1:
            style_list.append(
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), 
                 [colors.white, self.light_gray])
            )
        
        table.setStyle(TableStyle(style_list))
        return table
    
    def _save_chart_as_image(self, fig, temp_prefix='chart'):
        """Save matplotlib figure as BytesIO buffer for PDF inclusion"""
        if fig is None:
            return None
        
        # Save to BytesIO
        img_buffer = BytesIO()
        fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight', 
                facecolor='white', edgecolor='none')
        img_buffer.seek(0)
        
        # Close matplotlib figure
        plt.close(fig)
        
        return img_buffer  # Return buffer, not ImageReader
    
    def _add_chart_to_story(self, story, chart_fig, caption, width=5*inch):
        """Add chart with caption to story"""
        if chart_fig is None:
            return
        
        # Save figure to BytesIO buffer
        img_buffer = BytesIO()
        chart_fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight', 
                        facecolor='white', edgecolor='none')
        img_buffer.seek(0)
        
        # Calculate height maintaining aspect ratio
        # Get image dimensions using PIL
        from PIL import Image as PILImage
        pil_img = PILImage.open(img_buffer)
        img_width, img_height = pil_img.size
        aspect = img_height / float(img_width)
        height = width * aspect
        
        # Reset buffer position
        img_buffer.seek(0)
        
        # Create ReportLab Image directly from buffer
        chart_image = Image(img_buffer, width=width, height=height)
        
        # Add caption
        caption_para = Paragraph(
            f"<i>{caption}</i>", 
            self.styles['CorporateSmall']
        )
        
        story.append(chart_image)
        story.append(caption_para)
        story.append(Spacer(1, 0.15*inch))
        
        # Close matplotlib figure
        plt.close(chart_fig)
    
    
    
    def _create_cover_page(self):
        """Simple corporate cover page with signature blocks"""
        elements = []
        
        # Top spacing
        elements.append(Spacer(1, 0.5*inch))
        
        # Organization header
        elements.append(Paragraph(
            self.branding.get('organization_name', 'DSWD Analytics Report'),
            ParagraphStyle('OrgName', parent=self.styles['Normal'], 
                          fontSize=14, textColor=self.secondary_color,
                          alignment=TA_CENTER, fontName='Helvetica-Bold')
        ))
        elements.append(Spacer(1, 0.2*inch))
        
        elements.append(Paragraph(
            self.office_name,
            ParagraphStyle('Office', parent=self.styles['Normal'], 
                          fontSize=11, textColor=colors.HexColor('#666666'),
                          alignment=TA_CENTER, fontName='Helvetica')
        ))
        elements.append(Spacer(1, 0.8*inch))
        
        # Report Title
        elements.append(Paragraph(
            "<b>COMPREHENSIVE ANALYTICS REPORT</b>",
            self.styles['CorporateTitle']
        ))
        
        elements.append(Paragraph(
            "Applicant Assistance Analysis",
            ParagraphStyle('Subtitle', parent=self.styles['Normal'], 
                          fontSize=12, textColor=self.secondary_color,
                          alignment=TA_CENTER, fontName='Helvetica')
        ))
        elements.append(Spacer(1, 0.5*inch))

        location_filters = self._format_location_filter()
        
        # Metadata box      
        metadata = [
            [Paragraph("<b>Report Information</b>", self.styles['CorporateTableCell']), ""],
            [Paragraph("Document Reference:", self.styles['CorporateTableCell']), 
             Paragraph(self.doc_ref_number, self.styles['CorporateTableCell'])],
            [Paragraph("Report Period:", self.styles['CorporateTableCell']),
             Paragraph(f"{self.filters.get('start_date', 'All Records')} to {self.filters.get('end_date', 'Present')}", 
                      self.styles['CorporateTableCell'])],
        ]
        
        # Add each location filter as separate row
        for location_filter in location_filters:
            label, value = location_filter.split(":", 1)
            metadata.append([
                Paragraph(f"{label.strip()}:", self.styles['CorporateTableCell']),
                Paragraph(value.strip(), self.styles['CorporateTableCell'])
            ])
        
        metadata.extend([
            [Paragraph("Date Generated:", self.styles['CorporateTableCell']),
             Paragraph(datetime.now().strftime('%B %d, %Y'), self.styles['CorporateTableCell'])],
            [Paragraph("Effectivity Date:", self.styles['CorporateTableCell']),
             Paragraph(self.effectivity_date, self.styles['CorporateTableCell'])],
            [Paragraph("Classification:", self.styles['CorporateTableCell']),
             Paragraph("<b>OFFICIAL USE ONLY</b>", self.styles['CorporateTableCell'])],
            [Paragraph("Total Applicants:", self.styles['CorporateTableCell']),
             Paragraph(f"<b>{self.data['summary']['total_applicants']:,}</b>", self.styles['CorporateTableCell'])],
        ])

        table = self._create_table(metadata, col_widths=[2.5*inch, 3*inch], 
                                   has_header=False, zebra=False)
        elements.append(table)
        
        elements.append(Spacer(1, 0.6*inch))
        
        # Signature blocks
        sig_data = [
            [Paragraph("<b>Prepared by:</b>", self.styles['SignatureBlock']),
             Paragraph("<b>Reviewed by:</b>", self.styles['SignatureBlock']),
             Paragraph("<b>Approved by:</b>", self.styles['SignatureBlock'])],
            ["", "", ""],
            ["_____________________", "_____________________", "_____________________"],
            [Paragraph(self.prepared_by, self.styles['SignatureBlock']),
             Paragraph(self.reviewed_by, self.styles['SignatureBlock']),
             Paragraph(self.approved_by, self.styles['SignatureBlock'])],
            [Paragraph("<i>Data Analyst</i>", self.styles['CorporateSmall']),
             Paragraph("<i>Operations Manager</i>", self.styles['CorporateSmall']),
             Paragraph("<i>Director</i>", self.styles['CorporateSmall'])],
        ]
        
        sig_table = Table(sig_data, colWidths=[2*inch, 2*inch, 2*inch])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(sig_table)
        
        elements.append(Spacer(1, 0.3*inch))
        
        elements.append(PageBreak())
        return elements

    def _create_table_of_contents(self):
        """Auto-generated table of contents"""
        elements = []
        
        elements.append(Paragraph("TABLE OF CONTENTS", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.15*inch))
        
        toc_data = [
            [Paragraph("<b>Section</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Page</b>", self.styles['CorporateTableCell'])],
            [Paragraph("1. Executive Summary", self.styles['CorporateTableCell']), "3"],
            [Paragraph("2. Geographic Analysis", self.styles['CorporateTableCell']), "5"],
            [Paragraph("3. Demographics & Economics", self.styles['CorporateTableCell']), "7"],
            [Paragraph("4. Trends Analysis", self.styles['CorporateTableCell']), "9"],
            [Paragraph("5. Performance Metrics", self.styles['CorporateTableCell']), "11"],
            [Paragraph("6. Insights & Recommendations", self.styles['CorporateTableCell']), "13"],
            [Paragraph("Appendix A: List of Figures", self.styles['CorporateTableCell']), "15"],
            [Paragraph("Appendix B: List of Tables", self.styles['CorporateTableCell']), "16"],
            [Paragraph("Appendix C: Methodology", self.styles['CorporateTableCell']), "17"],
        ]
        
        table = self._create_table(toc_data, col_widths=[5*inch, 1*inch])
        elements.append(table)
        
        elements.append(PageBreak())
        return elements
    
    def _create_executive_summary(self):
        """Section 1: Executive Summary (matches Excel Sheet 1)"""
        elements = []
        
        elements.append(Paragraph("1. EXECUTIVE SUMMARY", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.1*inch))
        
        summary = self.data['summary']
        
        # KPI Summary Table
        elements.append(Paragraph("1.1 Key Performance Indicators", self.styles['SubSection']))
        
        kpi_data = [
            [Paragraph("<b>Metric</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Value</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Status/Comparison</b>", self.styles['CorporateTableCell'])],
            [Paragraph("Total Applicants Processed", self.styles['CorporateTableCell']),
             Paragraph(f"{summary['total_applicants']:,}", self.styles['CorporateTableCell']),
             Paragraph(f"Growth: {summary['growth_rate']:+.1f}%", self.styles['CorporateTableCell'])],
            [Paragraph("Average Processing Time", self.styles['CorporateTableCell']),
             Paragraph(f"{summary['avg_processing_minutes']:.1f} minutes", self.styles['CorporateTableCell']),
             Paragraph("Target: <10 min", self.styles['CorporateTableCell'])],
            [Paragraph("Most Common Assistance Type", self.styles['CorporateTableCell']),
             Paragraph(summary['most_common_type'], self.styles['CorporateTableCell']),
             Paragraph("Primary Service", self.styles['CorporateTableCell'])],
            [Paragraph("Top Service Location", self.styles['CorporateTableCell']),
             Paragraph(summary['top_barangay'], self.styles['CorporateTableCell']),
             Paragraph("Highest Volume", self.styles['CorporateTableCell'])],
            [Paragraph("This Month Applications", self.styles['CorporateTableCell']),
             Paragraph(f"{summary['this_month']:,}", self.styles['CorporateTableCell']),
             Paragraph(f"vs Last Month: {summary['last_month']:,}", self.styles['CorporateTableCell'])],
        ]
        
        table = self._create_table(kpi_data, col_widths=[2.3*inch, 1.7*inch, 2*inch])
        elements.append(table)
        elements.append(Spacer(1, 0.15*inch))
        
        # Executive Summary Narrative
        elements.append(Paragraph("1.2 Summary Overview", self.styles['SubSection']))
        elements.append(Paragraph(
            self.insights.get('executive_summary', ''),
            self.styles['CorporateBody']
        ))
        
        elements.append(PageBreak())
        return elements

    def _create_geographic_analysis(self):
        """Section 2: Geographic Analysis (matches Excel Sheet 2)"""
        elements = []
        
        elements.append(Paragraph("2. GEOGRAPHIC ANALYSIS", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.1*inch))
        
        geo = self.data['geographic']
        
        # 2.1 Top Barangays
        elements.append(Paragraph("2.1 Top Barangays by Application Volume", self.styles['SubSection']))
        
        if geo['top_barangays']:
            # Table with top 10
            barangay_data = [
                [Paragraph("<b>Rank</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Barangay</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Applications</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Percentage</b>", self.styles['CorporateTableCell'])]
            ]
            
            total = sum(item['count'] for item in geo['top_barangays'])
            for idx, item in enumerate(geo['top_barangays'][:10], start=1):
                percentage = (item['count'] / total * 100) if total > 0 else 0
                barangay_data.append([
                    Paragraph(str(idx), self.styles['CorporateTableCell']),
                    Paragraph(item['background_info__barangay__name'], self.styles['CorporateTableCell']),
                    Paragraph(f"{item['count']:,}", self.styles['CorporateTableCell']),
                    Paragraph(f"{percentage:.1f}%", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(barangay_data, col_widths=[0.6*inch, 2.5*inch, 1.3*inch, 1*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.1*inch))
        
        elements.append(Spacer(1, 0.15*inch))
        
        # 2.2 Applications by City
        elements.append(Paragraph("2.2 Applications by City", self.styles['SubSection']))
        
        if geo['by_city']:
            city_data = [
                [Paragraph("<b>City</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Applications</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Percentage</b>", self.styles['CorporateTableCell'])]
            ]
            
            total = sum(item['count'] for item in geo['by_city'])
            for item in geo['by_city']:
                percentage = (item['count'] / total * 100) if total > 0 else 0
                city_data.append([
                    Paragraph(item['background_info__barangay__city__name'], self.styles['CorporateTableCell']),
                    Paragraph(f"{item['count']:,}", self.styles['CorporateTableCell']),
                    Paragraph(f"{percentage:.1f}%", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(city_data, col_widths=[3*inch, 1.5*inch, 1.5*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.1*inch))
        
        elements.append(Spacer(1, 0.15*inch))
        
        # 2.3 Approval Rates by Location
        elements.append(Paragraph("2.3 Approval Rates by Location", self.styles['SubSection']))
        
        if geo['approval_by_location']:
            approval_data = [
                [Paragraph("<b>City</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Total</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Approved</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Approval Rate</b>", self.styles['CorporateTableCell'])]
            ]
            
            for item in geo['approval_by_location'][:10]:
                city = item['background_info__barangay__city__name']
                total = item['total']
                approved = item['approved']
                rate = (approved / total * 100) if total > 0 else 0
                
                approval_data.append([
                    Paragraph(city, self.styles['CorporateTableCell']),
                    Paragraph(str(total), self.styles['CorporateTableCell']),
                    Paragraph(str(approved), self.styles['CorporateTableCell']),
                    Paragraph(f"{rate:.1f}%", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(approval_data, col_widths=[2.5*inch, 1*inch, 1*inch, 1.5*inch])
            elements.append(table)
        
        # Geographic Insights
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("2.4 Geographic Insights", self.styles['SubSection']))
        
        if self.insights.get('geographic'):
            for insight in self.insights['geographic']:
                elements.append(Paragraph(
                    insight,
                    self.styles['CorporateBullet']
                ))
        
        elements.append(PageBreak())
        return elements
    
    def _create_demographics_economics(self):
        """Section 3: Demographics & Economics (matches Excel Sheet 3)"""
        elements = []
        
        elements.append(Paragraph("3. DEMOGRAPHICS & ECONOMICS", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.1*inch))
        
        demo = self.data['demographics']
        econ = self.data['economics']
        
        # ===== DEMOGRAPHICS =====
        elements.append(Paragraph("3.1 Demographic Analysis", self.styles['SubSection']))
        
        # 3.1.1 Gender Distribution
        elements.append(Paragraph("3.1.1 Gender Distribution", 
                                 ParagraphStyle('SubSubSection', parent=self.styles['CorporateBody'],
                                              fontSize=10, fontName='Helvetica-Bold',
                                              spaceAfter=6, spaceBefore=8)))
        
        if demo['by_gender']:
            gender_data = [
                [Paragraph("<b>Gender</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Count</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Percentage</b>", self.styles['CorporateTableCell'])]
            ]
            
            total = sum(item['count'] for item in demo['by_gender'])
            for item in demo['by_gender']:
                percentage = (item['count'] / total * 100) if total > 0 else 0
                gender_data.append([
                    Paragraph(item['background_info__sex'], self.styles['CorporateTableCell']),
                    Paragraph(f"{item['count']:,}", self.styles['CorporateTableCell']),
                    Paragraph(f"{percentage:.1f}%", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(gender_data, col_widths=[2*inch, 2*inch, 2*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # 3.1.2 Age Groups
        elements.append(Paragraph("3.1.2 Age Group Distribution", 
                                 ParagraphStyle('SubSubSection', parent=self.styles['CorporateBody'],
                                              fontSize=10, fontName='Helvetica-Bold',
                                              spaceAfter=6, spaceBefore=8)))
        
        if demo['age_groups']:
            age_data = [
                [Paragraph("<b>Age Group</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Count</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Percentage</b>", self.styles['CorporateTableCell'])]
            ]
            
            total = sum(demo['age_groups'].values())
            for group, count in demo['age_groups'].items():
                percentage = (count / total * 100) if total > 0 else 0
                age_data.append([
                    Paragraph(group, self.styles['CorporateTableCell']),
                    Paragraph(f"{count:,}", self.styles['CorporateTableCell']),
                    Paragraph(f"{percentage:.1f}%", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(age_data, col_widths=[2*inch, 2*inch, 2*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # 3.1.3 Civil Status
        elements.append(Paragraph("3.1.3 Civil Status Distribution", 
                                 ParagraphStyle('SubSubSection', parent=self.styles['CorporateBody'],
                                              fontSize=10, fontName='Helvetica-Bold',
                                              spaceAfter=6, spaceBefore=8)))
        
        if demo['by_civil_status']:
            civil_data = [
                [Paragraph("<b>Civil Status</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Count</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Percentage</b>", self.styles['CorporateTableCell'])]
            ]
            
            total = sum(item['count'] for item in demo['by_civil_status'])
            for item in demo['by_civil_status']:
                percentage = (item['count'] / total * 100) if total > 0 else 0
                civil_data.append([
                    Paragraph(item['background_info__civil_status'], self.styles['CorporateTableCell']),
                    Paragraph(f"{item['count']:,}", self.styles['CorporateTableCell']),
                    Paragraph(f"{percentage:.1f}%", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(civil_data, col_widths=[2*inch, 2*inch, 2*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # 3.1.4 Top Occupations
        elements.append(Paragraph("3.1.4 Top 10 Occupations", 
                                 ParagraphStyle('SubSubSection', parent=self.styles['CorporateBody'],
                                              fontSize=10, fontName='Helvetica-Bold',
                                              spaceAfter=6, spaceBefore=8)))
        
        if demo['by_occupation']:
            occ_data = [
                [Paragraph("<b>Occupation</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Count</b>", self.styles['CorporateTableCell'])]
            ]
            
            for item in demo['by_occupation'][:10]:
                occ_data.append([
                    Paragraph(item['background_info__occupation'], self.styles['CorporateTableCell']),
                    Paragraph(f"{item['count']:,}", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(occ_data, col_widths=[4*inch, 2*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # Demographic Insights
        elements.append(Paragraph("3.1.5 Demographic Insights", 
                                 ParagraphStyle('SubSubSection', parent=self.styles['CorporateBody'],
                                              fontSize=10, fontName='Helvetica-Bold',
                                              spaceAfter=6, spaceBefore=8)))
        
        if self.insights.get('demographic'):
            for insight in self.insights['demographic']:
                elements.append(Paragraph(insight, self.styles['CorporateBullet']))
        
        elements.append(Spacer(1, 0.2*inch))
        
        # ===== ECONOMICS =====
        elements.append(Paragraph("3.2 Economic Analysis", self.styles['SubSection']))
        
        # 3.2.1 Income Distribution
        elements.append(Paragraph("3.2.1 Monthly Income Distribution", 
                                 ParagraphStyle('SubSubSection', parent=self.styles['CorporateBody'],
                                              fontSize=10, fontName='Helvetica-Bold',
                                              spaceAfter=6, spaceBefore=8)))
        
        if econ['income_distribution']:
            income_data = [
                [Paragraph("<b>Income Range (PHP)</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Applicants</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Percentage</b>", self.styles['CorporateTableCell'])]
            ]
            
            total = sum(item['count'] for item in econ['income_distribution'])
            for item in econ['income_distribution']:
                percentage = (item['count'] / total * 100) if total > 0 else 0
                income_data.append([
                    Paragraph(item['range'], self.styles['CorporateTableCell']),
                    Paragraph(f"{item['count']:,}", self.styles['CorporateTableCell']),
                    Paragraph(f"{percentage:.1f}%", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(income_data, col_widths=[2.5*inch, 1.5*inch, 1.5*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # 3.2.2 Economic Indicators
        elements.append(Paragraph("3.2.2 Economic Indicators", 
                                 ParagraphStyle('SubSubSection', parent=self.styles['CorporateBody'],
                                              fontSize=10, fontName='Helvetica-Bold',
                                              spaceAfter=6, spaceBefore=8)))
        
        if econ['income_distribution']:
            # Calculate indicators
            below_poverty = econ['income_distribution'][0]['count'] if len(econ['income_distribution']) > 0 else 0
            middle_income = sum(item['count'] for item in econ['income_distribution'][2:5]) if len(econ['income_distribution']) >= 5 else 0
            above_average = sum(item['count'] for item in econ['income_distribution'][5:]) if len(econ['income_distribution']) > 5 else 0
            
            total_economic = below_poverty + middle_income + above_average
            
            indicators_data = [
                [Paragraph("<b>Indicator</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Count</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Percentage</b>", self.styles['CorporateTableCell'])]
            ]
            
            indicators = [
                ("Below Poverty Threshold (<10k)", below_poverty),
                ("Middle Income (20k-50k)", middle_income),
                ("Above Average (>50k)", above_average)
            ]
            
            for indicator, count in indicators:
                percentage = (count / total_economic * 100) if total_economic > 0 else 0
                indicators_data.append([
                    Paragraph(indicator, self.styles['CorporateTableCell']),
                    Paragraph(f"{count:,}", self.styles['CorporateTableCell']),
                    Paragraph(f"{percentage:.1f}%", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(indicators_data, col_widths=[3*inch, 1.5*inch, 1.5*inch])
            elements.append(table)
        
        elements.append(PageBreak())
        return elements

    def _create_trends_analysis(self):
        """Section 4: Trends Analysis (matches Excel Sheet 4)"""
        elements = []
        
        elements.append(Paragraph("4. TRENDS ANALYSIS", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.1*inch))
        
        trends = self.data['trends']
        
        # 4.1 Monthly Trends
        elements.append(Paragraph("4.1 Monthly Application Trends (Last 12 Months)", self.styles['SubSection']))
        
        if trends['monthly']:
            monthly_data = [
                [Paragraph("<b>Month</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Applications</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Change</b>", self.styles['CorporateTableCell'])]
            ]
            
            for idx, item in enumerate(trends['monthly']):
                month_str = item['month'].strftime('%B %Y') if item['month'] else 'N/A'
                count = item['count']
                
                # Calculate change from previous month
                if idx > 0:
                    prev_count = trends['monthly'][idx - 1]['count']
                    change = count - prev_count
                    change_pct = (change / prev_count * 100) if prev_count > 0 else 0
                    change_str = f"{change:+,} ({change_pct:+.1f}%)"
                else:
                    change_str = "—"
                
                monthly_data.append([
                    Paragraph(month_str, self.styles['CorporateTableCell']),
                    Paragraph(f"{count:,}", self.styles['CorporateTableCell']),
                    Paragraph(change_str, self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(monthly_data, col_widths=[2.5*inch, 1.5*inch, 2*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # 4.2 Yearly Trends
        elements.append(Paragraph("4.2 Yearly Application Trends", self.styles['SubSection']))
        
        if trends['yearly']:
            yearly_data = [
                [Paragraph("<b>Year</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Applications</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Year-over-Year Change</b>", self.styles['CorporateTableCell'])]
            ]
            
            for idx, item in enumerate(trends['yearly']):
                year = item['year']
                count = item['count']
                
                # Calculate YoY change
                if idx > 0:
                    prev_count = trends['yearly'][idx - 1]['count']
                    change = count - prev_count
                    change_pct = (change / prev_count * 100) if prev_count > 0 else 0
                    change_str = f"{change:+,} ({change_pct:+.1f}%)"
                else:
                    change_str = "—"
                
                yearly_data.append([
                    Paragraph(str(year), self.styles['CorporateTableCell']),
                    Paragraph(f"{count:,}", self.styles['CorporateTableCell']),
                    Paragraph(change_str, self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(yearly_data, col_widths=[2*inch, 2*inch, 2*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # 4.3 Assistance Type Distribution
        elements.append(Paragraph("4.3 Applications by Assistance Type", self.styles['SubSection']))
        
        if trends['by_assistance']:
            assist_data = [
                [Paragraph("<b>Assistance Type</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Applications</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Percentage</b>", self.styles['CorporateTableCell'])]
            ]
            
            total = sum(item['count'] for item in trends['by_assistance'])
            for item in trends['by_assistance']:
                percentage = (item['count'] / total * 100) if total > 0 else 0

                assistance_type = item.get('type_of_assistance') or 'Unknown Type'

                assist_data.append([
                    Paragraph(str(assistance_type), self.styles['CorporateTableCell']),
                    Paragraph(f"{item['count']:,}", self.styles['CorporateTableCell']),
                    Paragraph(f"{percentage:.1f}%", self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(assist_data, col_widths=[3*inch, 1.5*inch, 1.5*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # Trend Insights
        elements.append(Paragraph("4.4 Trend Insights", self.styles['SubSection']))
        
        if self.insights.get('trends'):
            for insight in self.insights['trends']:
                elements.append(Paragraph(insight, self.styles['CorporateBullet']))
        
        elements.append(PageBreak())
        return elements

    def _create_performance_metrics(self):
        """Section 5: Performance Metrics (matches Excel Sheet 5)"""
        elements = []
        
        elements.append(Paragraph("5. PERFORMANCE METRICS", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.1*inch))
        
        perf = self.data['performance']
        
        # 5.1 Staff Productivity
        elements.append(Paragraph("5.1 Staff Productivity (Top 10)", self.styles['SubSection']))
        
        if perf['staff_productivity']:
            staff_data = [
                [Paragraph("<b>Rank</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Staff Member</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Applications Processed</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Performance</b>", self.styles['CorporateTableCell'])]
            ]
            
            total_apps = sum(item['count'] for item in perf['staff_productivity'])
            avg_per_staff = total_apps / len(perf['staff_productivity']) if perf['staff_productivity'] else 0
            
            for idx, item in enumerate(perf['staff_productivity'][:10], start=1):
                count = item['count']
                performance = "Above Avg" if count > avg_per_staff else "Below Avg"

                username = item.get('staff__username') or 'Unknown Staff'
                
                staff_data.append([
                    Paragraph(str(idx), self.styles['CorporateTableCell']),
                    Paragraph(str(username), self.styles['CorporateTableCell']),
                    Paragraph(f"{count:,}", self.styles['CorporateTableCell']),
                    Paragraph(performance, self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(staff_data, col_widths=[0.6*inch, 2.5*inch, 1.8*inch, 1.5*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # 5.2 Processing Time by Assistance Type
        elements.append(Paragraph("5.2 Average Processing Time by Assistance Type", self.styles['SubSection']))
        
        if perf['processing_by_type']:
            proc_data = [
                [Paragraph("<b>Assistance Type</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Avg Time (min)</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Status</b>", self.styles['CorporateTableCell'])]
            ]
            
            for item in perf['processing_by_type']:
                avg_min = item.get('avg_minutes', 0)
                status = "Fast" if avg_min < 10 else "Needs Improvement"

                assistance_type = item.get('type_of_assistance') or 'Unknown Type'
                
                proc_data.append([
                    Paragraph(str(assistance_type), self.styles['CorporateTableCell']),
                    Paragraph(f"{avg_min:.1f}", self.styles['CorporateTableCell']),
                    Paragraph(status, self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(proc_data, col_widths=[3*inch, 1.5*inch, 1.5*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # 5.3 Activity Heatmap
        elements.append(Paragraph("5.3 Application Activity by Hour of Day", self.styles['SubSection']))
        
        if perf['activity_heatmap']:
            # Group by time periods for readability
            heat_data = [
                [Paragraph("<b>Time Period</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Hour Range</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Applications</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Activity Level</b>", self.styles['CorporateTableCell'])]
            ]
            
            max_count = max(item['count'] for item in perf['activity_heatmap'])
            
            # Group into periods
            periods = [
                ("Morning", "06:00 - 11:59", [6, 7, 8, 9, 10, 11]),
                ("Afternoon", "12:00 - 17:59", [12, 13, 14, 15, 16, 17]),
                ("Evening", "18:00 - 23:59", [18, 19, 20, 21, 22, 23]),
            ]
            
            for period_name, hour_range, hours in periods:
                period_count = sum(item['count'] for item in perf['activity_heatmap'] 
                                 if item['hour'] in hours)
                
                if period_count > 0:
                    activity_level = "High" if period_count > (max_count * 0.6) else "Medium" if period_count > (max_count * 0.3) else "Low"
                    
                    heat_data.append([
                        Paragraph(period_name, self.styles['CorporateTableCell']),
                        Paragraph(hour_range, self.styles['CorporateTableCell']),
                        Paragraph(f"{period_count:,}", self.styles['CorporateTableCell']),
                        Paragraph(activity_level, self.styles['CorporateTableCell'])
                    ])
            
            table = self._create_table(heat_data, col_widths=[1.5*inch, 1.8*inch, 1.5*inch, 1.5*inch])
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # 5.4 Performance Summary
        elements.append(Paragraph("5.4 Performance Summary", self.styles['SubSection']))
        
        avg_processing = self.data['summary']['avg_processing_minutes']
        total_staff = len(perf['staff_productivity']) if perf['staff_productivity'] else 0
        total_processed = sum(item['count'] for item in perf['staff_productivity']) if perf['staff_productivity'] else 0
        avg_per_staff = total_processed / total_staff if total_staff > 0 else 0
        
        peak_hour = max(perf['activity_heatmap'], key=lambda x: x['count'])['hour'] if perf['activity_heatmap'] else 0
        
        summary_data = [
            [Paragraph("<b>Metric</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Value</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Assessment</b>", self.styles['CorporateTableCell'])],
            [Paragraph("Average Processing Time", self.styles['CorporateTableCell']),
             Paragraph(f"{avg_processing:.1f} minutes", self.styles['CorporateTableCell']),
             Paragraph("Good" if avg_processing < 10 else "Needs Improvement", self.styles['CorporateTableCell'])],
            [Paragraph("Total Active Staff", self.styles['CorporateTableCell']),
             Paragraph(str(total_staff), self.styles['CorporateTableCell']),
             Paragraph("—", self.styles['CorporateTableCell'])],
            [Paragraph("Avg Applications per Staff", self.styles['CorporateTableCell']),
             Paragraph(f"{avg_per_staff:.1f}", self.styles['CorporateTableCell']),
             Paragraph("—", self.styles['CorporateTableCell'])],
            [Paragraph("Peak Activity Hour", self.styles['CorporateTableCell']),
             Paragraph(f"{peak_hour:02d}:00", self.styles['CorporateTableCell']),
             Paragraph("—", self.styles['CorporateTableCell'])],
        ]
        
        table = self._create_table(summary_data, col_widths=[2.5*inch, 1.8*inch, 1.7*inch])
        elements.append(table)
        elements.append(Spacer(1, 0.15*inch))
        
        # Performance Insights
        elements.append(Paragraph("5.5 Performance Insights", self.styles['SubSection']))
        
        if self.insights.get('performance'):
            for insight in self.insights['performance']:
                elements.append(Paragraph(insight, self.styles['CorporateBullet']))
        
        elements.append(PageBreak())
        return elements

    def _create_insights_recommendations(self):
        """Section 6: Insights & Recommendations (matches Excel Sheet 6)"""
        elements = []
        
        elements.append(Paragraph("6. INSIGHTS & RECOMMENDATIONS", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.1*inch))
        
        insights = self.insights
        
        # 6.1 Geographic Insights
        elements.append(Paragraph("6.1 Geographic Insights", self.styles['SubSection']))
        
        if insights.get('geographic'):
            for insight in insights['geographic']:
                elements.append(Paragraph(insight, self.styles['CorporateBullet']))
        else:
            elements.append(Paragraph("No significant geographic insights available.", self.styles['CorporateBody']))
        
        elements.append(Spacer(1, 0.15*inch))
        
        # 6.2 Demographic Insights
        elements.append(Paragraph("6.2 Demographic Insights", self.styles['SubSection']))
        
        if insights.get('demographic'):
            for insight in insights['demographic']:
                elements.append(Paragraph(insight, self.styles['CorporateBullet']))
        else:
            elements.append(Paragraph("No significant demographic insights available.", self.styles['CorporateBody']))
        
        elements.append(Spacer(1, 0.15*inch))
        
        # 6.3 Trend Insights
        elements.append(Paragraph("6.3 Trend Insights", self.styles['SubSection']))
        
        if insights.get('trends'):
            for insight in insights['trends']:
                elements.append(Paragraph(insight, self.styles['CorporateBullet']))
        else:
            elements.append(Paragraph("No significant trend insights available.", self.styles['CorporateBody']))
        
        elements.append(Spacer(1, 0.15*inch))
        
        # 6.4 Performance Insights
        elements.append(Paragraph("6.4 Performance Insights", self.styles['SubSection']))
        
        if insights.get('performance'):
            for insight in insights['performance']:
                elements.append(Paragraph(insight, self.styles['CorporateBullet']))
        else:
            elements.append(Paragraph("No significant performance insights available.", self.styles['CorporateBody']))
        
        elements.append(Spacer(1, 0.2*inch))
        
        # 6.5 Actionable Recommendations (Priority-based)
        elements.append(Paragraph("6.5 Actionable Recommendations", self.styles['SubSection']))
        
        if insights.get('recommendations'):
            # Create priority-based table
            rec_data = [
                [Paragraph("<b>Priority</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Recommendation</b>", self.styles['CorporateTableCell']),
                 Paragraph("<b>Impact</b>", self.styles['CorporateTableCell'])]
            ]
            
            # Assign priorities (customize logic as needed)
            priorities = ["HIGH", "HIGH", "MEDIUM", "LOW", "LOW"]
            impacts = ["High", "High", "Medium", "Low", "Low"]
            
            for idx, rec in enumerate(insights['recommendations']):
                priority = priorities[idx] if idx < len(priorities) else "LOW"
                impact = impacts[idx] if idx < len(impacts) else "Low"
                
                rec_data.append([
                    Paragraph(f"<b>{priority}</b>", self.styles['CorporateTableCell']),
                    Paragraph(rec.replace('•', '').strip(), self.styles['CorporateTableCell']),
                    Paragraph(impact, self.styles['CorporateTableCell'])
                ])
            
            table = self._create_table(rec_data, col_widths=[1*inch, 3.8*inch, 1*inch])
            elements.append(table)
        else:
            elements.append(Paragraph("No specific recommendations at this time.", self.styles['CorporateBody']))
        
        elements.append(PageBreak())
        return elements

    def _create_figures_appendix(self):
        """Appendix A: List of Figures"""
        elements = []
        
        elements.append(Paragraph("APPENDIX A: LIST OF FIGURES", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.15*inch))
        
        # Chart sections
        elements.append(Paragraph("Geographic Analysis Charts", self.styles['SubSection']))
        
        geo_charts = []
        if 'top_barangays' in self.charts:
            geo_charts.append("Figure 2.1: Top 10 Barangays by Application Volume")
        if 'by_city' in self.charts:
            geo_charts.append("Figure 2.2: Applications by City Distribution")
        
        if geo_charts:
            for chart in geo_charts:
                elements.append(Paragraph(f"• {chart}", self.styles['CorporateBullet']))
        else:
            elements.append(Paragraph("No geographic charts available.", self.styles['CorporateBody']))
        
        elements.append(Spacer(1, 0.15*inch))
        
        # Demographics charts
        elements.append(Paragraph("Demographics & Economics Charts", self.styles['SubSection']))
        
        demo_charts = []
        if 'by_gender' in self.charts:
            demo_charts.append("Figure 3.1: Gender Distribution")
        if 'age_groups' in self.charts:
            demo_charts.append("Figure 3.2: Age Group Distribution")
        if 'civil_status' in self.charts:
            demo_charts.append("Figure 3.3: Civil Status Distribution")
        if 'income' in self.charts:
            demo_charts.append("Figure 3.4: Monthly Income Distribution")
        
        if demo_charts:
            for chart in demo_charts:
                elements.append(Paragraph(f"• {chart}", self.styles['CorporateBullet']))
        else:
            elements.append(Paragraph("No demographic/economic charts available.", self.styles['CorporateBody']))
        
        elements.append(Spacer(1, 0.15*inch))
        
        # Trends charts
        elements.append(Paragraph("Trends Analysis Charts", self.styles['SubSection']))
        
        trend_charts = []
        if 'monthly_trend' in self.charts:
            trend_charts.append("Figure 4.1: Monthly Application Trends")
        if 'yearly_trend' in self.charts:
            trend_charts.append("Figure 4.2: Yearly Application Trends")
        if 'assistance_types' in self.charts:
            trend_charts.append("Figure 4.3: Applications by Assistance Type")
        
        if trend_charts:
            for chart in trend_charts:
                elements.append(Paragraph(f"• {chart}", self.styles['CorporateBullet']))
        else:
            elements.append(Paragraph("No trend charts available.", self.styles['CorporateBody']))
        
        elements.append(Spacer(1, 0.15*inch))
        
        # Performance charts
        elements.append(Paragraph("Performance Metrics Charts", self.styles['SubSection']))
        
        perf_charts = []
        if 'staff_productivity' in self.charts:
            perf_charts.append("Figure 5.1: Staff Productivity Comparison")
        if 'processing_time' in self.charts:
            perf_charts.append("Figure 5.2: Average Processing Time by Type")
        if 'activity_heatmap' in self.charts:
            perf_charts.append("Figure 5.3: Hourly Activity Pattern")
        
        if perf_charts:
            for chart in perf_charts:
                elements.append(Paragraph(f"• {chart}", self.styles['CorporateBullet']))
        else:
            elements.append(Paragraph("No performance charts available.", self.styles['CorporateBody']))
        
        elements.append(PageBreak())
        return elements

    def _create_tables_appendix(self):
        """Appendix B: List of Tables"""
        elements = []
        
        elements.append(Paragraph("APPENDIX B: LIST OF TABLES", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.15*inch))
        
        tables_list = [
            ("Table 1.1", "Key Performance Indicators"),
            ("Table 2.1", "Top 10 Barangays by Application Volume"),
            ("Table 2.2", "Applications by City"),
            ("Table 2.3", "Approval Rates by Location"),
            ("Table 3.1", "Gender Distribution"),
            ("Table 3.2", "Age Group Distribution"),
            ("Table 3.3", "Civil Status Distribution"),
            ("Table 3.4", "Top 10 Occupations"),
            ("Table 3.5", "Monthly Income Distribution"),
            ("Table 3.6", "Economic Indicators"),
            ("Table 4.1", "Monthly Application Trends"),
            ("Table 4.2", "Yearly Application Trends"),
            ("Table 4.3", "Applications by Assistance Type"),
            ("Table 5.1", "Staff Productivity Rankings"),
            ("Table 5.2", "Average Processing Time by Assistance Type"),
            ("Table 5.3", "Application Activity by Time Period"),
            ("Table 5.4", "Performance Summary Metrics"),
            ("Table 6.1", "Actionable Recommendations with Priority Levels"),
        ]
        
        table_data = [
            [Paragraph("<b>Table Number</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Table Title</b>", self.styles['CorporateTableCell'])]
        ]
        
        for table_num, table_title in tables_list:
            table_data.append([
                Paragraph(table_num, self.styles['CorporateTableCell']),
                Paragraph(table_title, self.styles['CorporateTableCell'])
            ])
        
        table = self._create_table(table_data, col_widths=[1.5*inch, 4.5*inch])
        elements.append(table)
        
        elements.append(PageBreak())
        return elements

    def _create_methodology_appendix(self):
        """Appendix C: Data Sources & Methodology"""
        elements = []
        
        elements.append(Paragraph("APPENDIX C: METHODOLOGY & DATA SOURCES", self.styles['SectionHeader']))
        elements.append(Spacer(1, 0.15*inch))
        
        # Data Sources
        elements.append(Paragraph("C.1 Data Sources", self.styles['SubSection']))
        
        source_text = f"""
This report is based on data extracted from the DSWD Applicant Management System. 
All data has been anonymized and aggregated to protect individual privacy in compliance 
with the Data Privacy Act of 2012 (RA 10173).
        """.strip()
        
        elements.append(Paragraph(source_text, self.styles['CorporateBody']))
        elements.append(Spacer(1, 0.1*inch))
        
        # Analysis Period
        elements.append(Paragraph("C.2 Analysis Period", self.styles['SubSection']))
        
        period_data = [
            [Paragraph("<b>Parameter</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Value</b>", self.styles['CorporateTableCell'])],
            [Paragraph("Start Date", self.styles['CorporateTableCell']),
             Paragraph(str(self.filters.get('start_date', 'All Records')), self.styles['CorporateTableCell'])],
            [Paragraph("End Date", self.styles['CorporateTableCell']),
             Paragraph(str(self.filters.get('end_date', 'Present')), self.styles['CorporateTableCell'])],
            [Paragraph("Total Records Analyzed", self.styles['CorporateTableCell']),
             Paragraph(f"{self.data['summary']['total_applicants']:,}", self.styles['CorporateTableCell'])],
            [Paragraph("Report Generated", self.styles['CorporateTableCell']),
             Paragraph(datetime.now().strftime('%B %d, %Y at %I:%M %p'), self.styles['CorporateTableCell'])],
            [Paragraph("Generated By", self.styles['CorporateTableCell']),
             Paragraph(self.prepared_by, self.styles['CorporateTableCell'])],
        ]
        
        table = self._create_table(period_data, col_widths=[2.5*inch, 3.5*inch], zebra=False)
        elements.append(table)
        elements.append(Spacer(1, 0.15*inch))
        
        # Filters Applied
        elements.append(Paragraph("C.3 Filters Applied", self.styles['SubSection']))
        
        filters_applied = []
        
        if self.filters.get('cities'):
            filters_applied.append(f"Cities: {', '.join(self.filters['cities'])}")
        else:
            filters_applied.append("Cities: All cities included")
        
        if self.filters.get('barangays'):
            filters_applied.append(f"Barangays: {', '.join(self.filters['barangays'][:5])}{'...' if len(self.filters['barangays']) > 5 else ''}")
        else:
            filters_applied.append("Barangays: All barangays included")
        
        if self.filters.get('assistance_types'):
            filters_applied.append(f"Assistance Types: {', '.join(self.filters['assistance_types'])}")
        else:
            filters_applied.append("Assistance Types: All types included")
        
        for filter_text in filters_applied:
            elements.append(Paragraph(f"• {filter_text}", self.styles['CorporateBullet']))
        
        elements.append(Spacer(1, 0.15*inch))
        
        # Methodology
        elements.append(Paragraph("C.4 Statistical Methodology", self.styles['SubSection']))
        
        methodology_text = """
The analysis employs descriptive statistics and trend analysis using Python-based analytics 
with Django ORM queries. Key metrics include:

- <b>Frequency Distribution:</b> Count and percentage calculations for categorical variables
- <b>Central Tendency:</b> Average (mean) calculations for processing times
- <b>Trend Analysis:</b> Month-over-month and year-over-year growth rates
- <b>Comparative Analysis:</b> Geographic and demographic segmentation
- <b>Performance Metrics:</b> Staff productivity and processing efficiency indicators

All calculations exclude null values and archived records. Age calculations are based on 
birthdates as of the report generation date.
        """.strip()
        
        elements.append(Paragraph(methodology_text, self.styles['CorporateBody']))
        elements.append(Spacer(1, 0.15*inch))
        
        # Data Quality
        elements.append(Paragraph("C.5 Data Quality & Limitations", self.styles['SubSection']))
        
        quality_text = """
<b>Data Quality Assurance:</b>
- All data validated against system integrity constraints
- Duplicate records removed during processing
- Incomplete records excluded from specific calculations

<b>Known Limitations:</b>
- Processing time calculations limited to records with complete timestamps
- Income data may be self-reported and subject to verification
- Historical data prior to system implementation may be incomplete
- Approval rates subject to ongoing case processing and updates
        """.strip()
        
        elements.append(Paragraph(quality_text, self.styles['CorporateBody']))
        elements.append(Spacer(1, 0.15*inch))
        
        # Version Control
        elements.append(Paragraph("C.6 Version Control", self.styles['SubSection']))
        
        version_data = [
            [Paragraph("<b>Version</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Date</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Changes</b>", self.styles['CorporateTableCell']),
             Paragraph("<b>Author</b>", self.styles['CorporateTableCell'])],
            [Paragraph("1.0", self.styles['CorporateTableCell']),
             Paragraph(datetime.now().strftime('%Y-%m-%d'), self.styles['CorporateTableCell']),
             Paragraph("Initial Report Release", self.styles['CorporateTableCell']),
             Paragraph(self.prepared_by, self.styles['CorporateTableCell'])],
        ]
        
        table = self._create_table(version_data, col_widths=[0.8*inch, 1.2*inch, 2.5*inch, 1.5*inch])
        elements.append(table)
        
        elements.append(PageBreak())
        return elements

    def _create_charts_section(self):
        """Create separate pages for all charts"""
        elements = []
        
        # Only add charts if they exist
        chart_sections = [
            ("GEOGRAPHIC ANALYSIS CHARTS", [
                ('top_barangays', "Figure 2.1: Top 10 Barangays by Application Volume"),
                ('by_city', "Figure 2.2: Applications by City Distribution"),
            ]),
            ("DEMOGRAPHICS & ECONOMICS CHARTS", [
                ('by_gender', "Figure 3.1: Gender Distribution"),
                ('age_groups', "Figure 3.2: Age Group Distribution"),
                ('civil_status', "Figure 3.3: Civil Status Distribution"),
                ('income', "Figure 3.4: Monthly Income Distribution"),
            ]),
            ("TRENDS ANALYSIS CHARTS", [
                ('monthly_trend', "Figure 4.1: Monthly Application Trends"),
                ('yearly_trend', "Figure 4.2: Yearly Application Trends"),
                ('assistance_types', "Figure 4.3: Applications by Assistance Type"),
                ('assistance_over_time', "Figure 4.4: Assistance Types Over Time"),
            ]),
            ("PERFORMANCE METRICS CHARTS", [
                ('staff_productivity', "Figure 5.1: Staff Productivity Comparison"),
                ('processing_time', "Figure 5.2: Average Processing Time by Type"),
                ('activity_heatmap', "Figure 5.3: Hourly Activity Pattern"),
            ]),
        ]
        
        for section_title, charts_list in chart_sections:
            section_has_charts = any(chart_key in self.charts for chart_key, _ in charts_list)
            
            if section_has_charts:
                elements.append(Paragraph(section_title, self.styles['SectionHeader']))
                elements.append(Spacer(1, 0.15*inch))
                
                for chart_key, caption in charts_list:
                    if chart_key in self.charts:
                        self._add_chart_to_story(elements, self.charts[chart_key], 
                                                caption, width=6*inch)
                        elements.append(Spacer(1, 0.2*inch))
                
                elements.append(PageBreak())
        
        return elements

    def generate(self, output_path):
        """Generate complete PDF report matching 6-sheet Excel structure"""
        # Use A4 for international, Letter for US
        page_size = A4  # or letter
        
        doc = SimpleDocTemplate(
            output_path,
            pagesize=page_size,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.9*inch,   # Space for header
            bottomMargin=0.8*inch  # Space for footer
        )
        
        story = []
        
        # Build document structure (matches 6 Excel sheets)
        story.extend(self._create_cover_page())                    # Cover
        story.extend(self._create_table_of_contents())            # TOC
        story.extend(self._create_executive_summary())            # Sheet 1
        story.extend(self._create_geographic_analysis())          # Sheet 2
        story.extend(self._create_demographics_economics())       # Sheet 3
        story.extend(self._create_trends_analysis())              # Sheet 4
        story.extend(self._create_performance_metrics())          # Sheet 5
        story.extend(self._create_insights_recommendations())     # Sheet 6
        
        # Appendices
        story.extend(self._create_figures_appendix())             # Appendix A
        story.extend(self._create_tables_appendix())              # Appendix B
        story.extend(self._create_methodology_appendix())         # Appendix C
        
        # Charts section (separate pages, flexible sizing)
        story.extend(self._create_charts_section())
        
        # Build PDF with headers and footers
        doc.build(story, 
                 onFirstPage=self._corporate_header_footer, 
                 onLaterPages=self._corporate_header_footer)
        
        return True


class ExcelReportGenerator:
    """Enhanced Excel with 6 professional sheets for analytics reporting"""
    
    def __init__(self, data, insights, charts, branding, filters):
        self.data = data
        self.insights = insights
        self.charts = charts
        self.branding = branding
        self.filters = filters
        self.wb = Workbook()
        
        # Professional colors
        self.primary_color = self.branding.get('primary_color', '0066cc').replace('#', '')
        self.secondary_color = '003366'
        self.accent_color = '00cc66'
        self.light_blue = 'e3f2fd'
        self.light_gray = 'f5f5f5'
        
        # Define all styles
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup professional Excel styles"""
        # Fonts
        self.title_font = Font(name='Calibri', size=18, bold=True, color=self.primary_color)
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.subheader_font = Font(name='Calibri', size=11, bold=True, color=self.primary_color)
        self.body_font = Font(name='Calibri', size=10)
        self.metric_font = Font(name='Calibri', size=28, bold=True, color=self.primary_color)
        
        # Fills
        self.header_fill = PatternFill(start_color=self.primary_color, end_color=self.primary_color, fill_type='solid')
        self.light_fill = PatternFill(start_color=self.light_blue, end_color=self.light_blue, fill_type='solid')
        self.accent_fill = PatternFill(start_color=self.accent_color, end_color=self.accent_color, fill_type='solid')
        self.gray_fill = PatternFill(start_color=self.light_gray, end_color=self.light_gray, fill_type='solid')
        
        # Borders
        self.thin_border = Border(
            left=Side(style='thin', color='BDBDBD'),
            right=Side(style='thin', color='BDBDBD'),
            top=Side(style='thin', color='BDBDBD'),
            bottom=Side(style='thin', color='BDBDBD')
        )
        self.thick_border = Border(
            left=Side(style='medium', color=self.primary_color),
            right=Side(style='medium', color=self.primary_color),
            top=Side(style='medium', color=self.primary_color),
            bottom=Side(style='medium', color=self.primary_color)
        )
        
        # Alignments
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.right_align = Alignment(horizontal='right', vertical='center')


    def _add_report_header(self, ws, title):
        """Add professional header to each sheet"""
        # Merge cells for title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = self.branding.get('organization_name', 'DSWD Analytics Report')
        title_cell.font = Font(name='Calibri', size=14, bold=True, color=self.secondary_color)
        title_cell.alignment = self.center_align
        title_cell.fill = self.gray_fill
        
        # Sheet title
        ws.merge_cells('A2:H2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = title
        subtitle_cell.font = self.title_font
        subtitle_cell.alignment = self.center_align
        
        # Date range
        ws.merge_cells('A3:H3')
        date_cell = ws['A3']
        start = self.filters.get('start_date', 'All Records')
        end = self.filters.get('end_date', 'Present')
        date_cell.value = f"Report Period: {start} to {end}"
        date_cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
        date_cell.alignment = self.center_align

        current_row = 4 

        city = self.filters.get('city')
        cities = self.filters.get('cities', [])
        barangays = self.filters.get('barangays', [])
        assistance_types = self.filters.get('assistance_types', [])
        
        filter_info = []
        
        # City info
        if city:
            filter_info.append(f"City: {city}")
        elif cities:
            if len(cities) <= 3:
                filter_info.append(f"Cities: {', '.join(cities)}")
            else:
                filter_info.append(f"Cities: {', '.join(cities[:3])} +{len(cities)-3} more")
        else:
            filter_info.append("Location: All Cities")
        
        # Barangay info
        if barangays:
            if len(barangays) <= 3:
                filter_info.append(f"Barangays: {', '.join(barangays)}")
            else:
                filter_info.append(f"Barangays: {', '.join(barangays[:3])} +{len(barangays)-3} more")
        
        # Assistance types info
        if assistance_types:
            if len(assistance_types) <= 3:
                filter_info.append(f"Types: {', '.join(assistance_types)}")
            else:
                filter_info.append(f"Types: {', '.join(assistance_types[:2])} +{len(assistance_types)-2} more")
        else:
            filter_info.append("Types: All Types")
        
        # Add filter info
        if filter_info:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            filter_cell = ws[f'A{current_row}']
            filter_cell.value = " | ".join(filter_info)
            filter_cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
            filter_cell.alignment = self.center_align
            current_row += 1
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[current_row - 1].height = 20
        
        return current_row + 1
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 20
        
        return 5  # Return next available row
    
    def _create_kpi_card(self, ws, row, col, title, value, color='primary'):
        """Create a KPI card visual"""
        # Determine color
        if color == 'primary':
            card_color = self.primary_color
        elif color == 'accent':
            card_color = self.accent_color
        elif color == 'red':
            card_color = 'f44336'
        else:
            card_color = color
        
        # Title cell
        title_cell = ws.cell(row=row, column=col)
        title_cell.value = title
        title_cell.font = Font(name='Calibri', size=10, bold=True, color='666666')
        title_cell.alignment = self.center_align
        title_cell.fill = self.gray_fill
        title_cell.border = self.thin_border
        
        # Value cell
        value_cell = ws.cell(row=row + 1, column=col)
        value_cell.value = value
        value_cell.font = Font(name='Calibri', size=20, bold=True, color=card_color)
        value_cell.alignment = self.center_align
        value_cell.border = self.thin_border
        
        # Merge if needed for wider cards
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        
        ws.row_dimensions[row + 1].height = 40
    
    def _create_data_table(self, ws, start_row, headers, data, col_widths=None):
        """Create a professional data table"""
        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # Add data with alternating row colors
        for row_idx, row_data in enumerate(data, start=start_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.thin_border
                cell.alignment = self.left_align if col_idx == 1 else self.right_align
                
                # Zebra striping
                if (row_idx - start_row) % 2 == 0:
                    cell.fill = self.light_fill
        
        # Set column widths
        if col_widths:
            for col_idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        else:
            # Auto-fit columns
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Add conditional formatting for numeric columns (if applicable)
        if len(data) > 0 and len(data[0]) >= 2:
            try:
                ws.conditional_formatting.add(
                    f'B{start_row + 1}:B{start_row + len(data)}',
                    ColorScaleRule(
                        start_type='min', start_color='FFFFFF',
                        mid_type='percentile', mid_value=50, mid_color=self.light_blue,
                        end_type='max', end_color=self.primary_color
                    )
                )
            except:
                pass
        
        return start_row + len(data) + 2  # Return next available row
    
    def _create_dashboard_sheet(self):
        """NEW: Executive dashboard with KPIs"""
        ws = self.wb.active
        ws.title = "📊 Dashboard"
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = self.branding.get('organization_name', 'DSWD Quezon Province')
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # KPI Cards
        row = 7
        summary = self.data['summary']
        
        # KPI 1: Total Applicants
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "Total Applicants"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.light_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'A{row+1}:B{row+1}')
        ws[f'A{row+1}'] = summary['total_applicants']
        ws[f'A{row+1}'].font = Font(size=32, bold=True, color=self.primary_color)
        ws[f'A{row+1}'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[row+1].height = 50
        
        # KPI 2: Processing Time
        ws.merge_cells(f'D{row}:E{row}')
        ws[f'D{row}'] = "Avg Processing Time"
        ws[f'D{row}'].font = self.subheader_font
        ws[f'D{row}'].fill = self.light_fill
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'D{row+1}:E{row+1}')
        ws[f'D{row+1}'] = f"{summary['avg_processing_minutes']:.1f} min"
        ws[f'D{row+1}'].font = Font(size=28, bold=True, color='00cc66')
        ws[f'D{row+1}'].alignment = Alignment(horizontal='center')
        
        # KPI 3: Growth
        ws.merge_cells(f'G{row}:H{row}')
        ws[f'G{row}'] = "Monthly Growth"
        ws[f'G{row}'].font = self.subheader_font
        ws[f'G{row}'].fill = self.light_fill
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'G{row+1}:H{row+1}')
        growth_symbol = "↑" if summary['growth_rate'] > 0 else "↓"
        ws[f'G{row+1}'] = f"{growth_symbol} {abs(summary['growth_rate']):.1f}%"
        growth_color = '00cc66' if summary['growth_rate'] > 0 else 'f44336'
        ws[f'G{row+1}'].font = Font(size=28, bold=True, color=growth_color)
        ws[f'G{row+1}'].alignment = Alignment(horizontal='center')
    
    def _add_data_sheet_with_chart(self, sheet_name, data, headers, chart_type='bar'):
        """NEW: Add sheet with embedded chart"""
        ws = self.wb.create_sheet(sheet_name)
        
        # Add title
        ws.merge_cells('A1:D1')
        ws['A1'] = sheet_name
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border
        
        # Add data
        for row_idx, row_data in enumerate(data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.thin_border
                if row_idx % 2 == 0:
                    cell.fill = self.light_fill
        
        # Freeze panes
        ws.freeze_panes = 'A4'
        
        # Add filter
        ws.auto_filter.ref = f'A3:{get_column_letter(len(headers))}{len(data) + 3}'
        
        # Add embedded chart
        if len(data) > 0:
            if chart_type == 'bar':
                chart = BarChart()
                chart.title = sheet_name
                data_ref = Reference(ws, min_col=2, min_row=3, max_row=min(len(data) + 3, 20))
                cats_ref = Reference(ws, min_col=1, min_row=4, max_row=min(len(data) + 3, 20))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws.add_chart(chart, f'{get_column_letter(len(headers) + 2)}3')
        
        # Conditional formatting for numeric columns
        if len(data) > 0 and isinstance(data[0][1], (int, float)):
            ws.conditional_formatting.add(
                f'B4:B{len(data) + 3}',
                ColorScaleRule(
                    start_type='min', start_color='FFFFFF',
                    mid_type='percentile', mid_value=50, mid_color=self.light_blue,
                    end_type='max', end_color=self.primary_color
                )
            )
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 3, 60)

    def _add_data_sheet(self, sheet_name, data, headers):
        """Generic method to add data sheet"""
        ws = self.wb.create_sheet(sheet_name)
        
        # Add headers
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_data in data:
            ws.append(row_data)
        
        # Apply borders to all data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = self.thin_border
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_summary_sheet(self):
        """Sheet 1: Executive Summary with KPIs and paragraph"""
        ws = self.wb.active
        ws.title = "1. Summary"
        
        # Add header
        current_row = self._add_report_header(ws, "EXECUTIVE SUMMARY")
        
        summary = self.data['summary']
        
        # KPI Cards Row 1
        self._create_kpi_card(ws, current_row, 1, "Total Applicants", f"{summary['total_applicants']:,}", 'primary')
        self._create_kpi_card(ws, current_row, 3, "Avg Processing Time", f"{summary['avg_processing_minutes']:.1f} min", 'accent')
        self._create_kpi_card(ws, current_row, 5, "This Month", f"{summary['this_month']:,}", 'primary')
        self._create_kpi_card(ws, current_row, 7, "Last Month", f"{summary['last_month']:,}", 'primary')
        
        current_row += 3
        
        # KPI Cards Row 2
        growth_color = 'accent' if summary['growth_rate'] > 0 else 'red'
        growth_symbol = "↑" if summary['growth_rate'] > 0 else "↓"
        self._create_kpi_card(ws, current_row, 1, "Monthly Growth", 
                             f"{growth_symbol} {abs(summary['growth_rate']):.1f}%", growth_color)
        
        # Most common type
        ws.merge_cells(f'C{current_row}:D{current_row+1}')
        type_cell = ws.cell(row=current_row, column=3)
        type_cell.value = f"Most Common:\n{summary['most_common_type']}"
        type_cell.font = Font(name='Calibri', size=11, bold=True)
        type_cell.alignment = self.center_align
        type_cell.fill = self.light_fill
        type_cell.border = self.thin_border
        
        # Top barangay
        ws.merge_cells(f'E{current_row}:F{current_row+1}')
        barangay_cell = ws.cell(row=current_row, column=5)
        barangay_cell.value = f"Top Barangay:\n{summary['top_barangay']}"
        barangay_cell.font = Font(name='Calibri', size=11, bold=True)
        barangay_cell.alignment = self.center_align
        barangay_cell.fill = self.light_fill
        barangay_cell.border = self.thin_border
        
        current_row += 4
        
        # Executive Summary Paragraph
        ws.merge_cells(f'A{current_row}:H{current_row}')
        para_title = ws.cell(row=current_row, column=1)
        para_title.value = "Executive Summary"
        para_title.font = self.subheader_font
        para_title.alignment = self.left_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row+6}')
        para_cell = ws.cell(row=current_row, column=1)
        para_cell.value = self.insights.get('executive_summary', '')
        para_cell.font = self.body_font
        para_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        para_cell.border = self.thin_border
        
        ws.row_dimensions[current_row].height = 150
        
        # Set column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_geographic_sheet(self):
        """Sheet 2: Geographic Analysis - Combined view"""
        ws = self.wb.create_sheet("2. Geographic")
        
        current_row = self._add_report_header(ws, "GEOGRAPHIC ANALYSIS")
        
        geo = self.data['geographic']
        
        # Section 1: Top Barangays
        ws.merge_cells(f'A{current_row}:D{current_row}')
        section1 = ws.cell(row=current_row, column=1)
        section1.value = "Top 10 Barangays by Application Volume"
        section1.font = self.subheader_font
        section1.alignment = self.left_align
        
        current_row += 1
        
        if geo['top_barangays']:
            barangay_data = [[item['background_info__barangay__name'], item['count']] 
                            for item in geo['top_barangays'][:10]]
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Barangay', 'Applications'], 
                                                  barangay_data, [30, 15])
            
            # Add bar chart
            chart = BarChart()
            chart.title = "Top Barangays"
            chart.style = 10
            chart.height = 10
            chart.width = 20
            
            data_ref = Reference(ws, min_col=2, min_row=current_row-len(barangay_data)-1, 
                               max_row=current_row-2)
            cats_ref = Reference(ws, min_col=1, min_row=current_row-len(barangay_data), 
                               max_row=current_row-2)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            ws.add_chart(chart, f'F{current_row-len(barangay_data)-1}')
        
        current_row += 2
        
        # Section 2: Applications by City
        ws.merge_cells(f'A{current_row}:D{current_row}')
        section2 = ws.cell(row=current_row, column=1)
        section2.value = "Applications by City"
        section2.font = self.subheader_font
        section2.alignment = self.left_align
        
        current_row += 1
        
        if geo['by_city']:
            city_data = [[item['background_info__barangay__city__name'], item['count']] 
                        for item in geo['by_city']]
            current_row = self._create_data_table(ws, current_row, 
                                                  ['City', 'Applications'], 
                                                  city_data, [30, 15])
            
            # Add pie chart
            pie = PieChart()
            pie.title = "City Distribution"
            pie.height = 10
            pie.width = 20
            
            data_ref = Reference(ws, min_col=2, min_row=current_row-len(city_data)-1, 
                               max_row=current_row-2)
            cats_ref = Reference(ws, min_col=1, min_row=current_row-len(city_data), 
                               max_row=current_row-2)
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)
            
            ws.add_chart(pie, f'F{current_row-len(city_data)-1}')
        
        current_row += 2
        
        # Section 3: Approval Rates by Location
        ws.merge_cells(f'A{current_row}:D{current_row}')
        section3 = ws.cell(row=current_row, column=1)
        section3.value = "Approval Rates by Location"
        section3.font = self.subheader_font
        section3.alignment = self.left_align
        
        current_row += 1
        
        if geo['approval_by_location']:
            approval_data = []
            for item in geo['approval_by_location']:
                city = item['background_info__barangay__city__name']
                total = item['total']
                approved = item['approved']
                rate = (approved / total * 100) if total > 0 else 0
                approval_data.append([city, total, approved, f"{rate:.1f}%"])
            
            current_row = self._create_data_table(ws, current_row,
                                                  ['City', 'Total', 'Approved', 'Rate'],
                                                  approval_data, [25, 12, 12, 12])   
    
    def _create_demographics_economics_sheet(self):
        """Sheet 3: Demographics & Economics - Combined"""
        ws = self.wb.create_sheet("3. Demographics & Economics")
        
        current_row = self._add_report_header(ws, "DEMOGRAPHICS & ECONOMICS")
        
        demo = self.data['demographics']
        econ = self.data['economics']
        
        # ===== DEMOGRAPHICS SECTION =====
        ws.merge_cells(f'A{current_row}:H{current_row}')
        demo_header = ws.cell(row=current_row, column=1)
        demo_header.value = "DEMOGRAPHIC ANALYSIS"
        demo_header.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        demo_header.fill = self.header_fill
        demo_header.alignment = self.center_align
        
        current_row += 2
        
        # Gender Distribution
        ws.merge_cells(f'A{current_row}:D{current_row}')
        gender_title = ws.cell(row=current_row, column=1)
        gender_title.value = "Gender Distribution"
        gender_title.font = self.subheader_font
        
        current_row += 1
        
        if demo['by_gender']:
            gender_data = [[item['background_info__sex'], item['count']] 
                          for item in demo['by_gender']]
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Gender', 'Count'], 
                                                  gender_data, [20, 15])
            
            # Pie chart
            pie = PieChart()
            pie.title = "Gender Distribution"
            pie.height = 8
            pie.width = 12
            data_ref = Reference(ws, min_col=2, min_row=current_row-len(gender_data)-1, max_row=current_row-2)
            cats_ref = Reference(ws, min_col=1, min_row=current_row-len(gender_data), max_row=current_row-2)
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)
            ws.add_chart(pie, f'F{current_row-len(gender_data)-1}')
        
        current_row += 2
        
        # Age Groups
        ws.merge_cells(f'A{current_row}:D{current_row}')
        age_title = ws.cell(row=current_row, column=1)
        age_title.value = "Age Group Distribution"
        age_title.font = self.subheader_font
        
        current_row += 1
        
        if demo['age_groups']:
            age_data = [[group, count] for group, count in demo['age_groups'].items()]
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Age Group', 'Count'], 
                                                  age_data, [20, 15])
            
            # Bar chart
            chart = BarChart()
            chart.title = "Age Distribution"
            chart.height = 8
            chart.width = 12
            data_ref = Reference(ws, min_col=2, min_row=current_row-len(age_data)-1, max_row=current_row-2)
            cats_ref = Reference(ws, min_col=1, min_row=current_row-len(age_data), max_row=current_row-2)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, f'F{current_row-len(age_data)-1}')
        
        current_row += 2
        
        # Civil Status
        ws.merge_cells(f'A{current_row}:D{current_row}')
        civil_title = ws.cell(row=current_row, column=1)
        civil_title.value = "Civil Status Distribution"
        civil_title.font = self.subheader_font
        
        current_row += 1
        
        if demo['by_civil_status']:
            civil_data = [[item['background_info__civil_status'], item['count']] 
                         for item in demo['by_civil_status']]
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Civil Status', 'Count'], 
                                                  civil_data, [20, 15])
        
        current_row += 2
        
        # Top Occupations
        ws.merge_cells(f'A{current_row}:D{current_row}')
        occ_title = ws.cell(row=current_row, column=1)
        occ_title.value = "Top 10 Occupations"
        occ_title.font = self.subheader_font
        
        current_row += 1
        
        if demo['by_occupation']:
            occ_data = [[item['background_info__occupation'], item['count']] 
                       for item in demo['by_occupation'][:10]]
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Occupation', 'Count'], 
                                                  occ_data, [30, 15])
        
        current_row += 3
        
        # ===== ECONOMICS SECTION =====
        ws.merge_cells(f'A{current_row}:H{current_row}')
        econ_header = ws.cell(row=current_row, column=1)
        econ_header.value = "ECONOMIC ANALYSIS"
        econ_header.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        econ_header.fill = PatternFill(start_color=self.accent_color, end_color=self.accent_color, fill_type='solid')
        econ_header.alignment = self.center_align
        
        current_row += 2
        
        # Income Distribution
        ws.merge_cells(f'A{current_row}:D{current_row}')
        income_title = ws.cell(row=current_row, column=1)
        income_title.value = "Monthly Income Distribution"
        income_title.font = self.subheader_font
        
        current_row += 1
        
        if econ['income_distribution']:
            income_data = [[item['range'], item['count']] 
                          for item in econ['income_distribution']]
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Income Range (PHP)', 'Applicants'], 
                                                  income_data, [25, 15])
            
            # Bar chart
            chart = BarChart()
            chart.title = "Income Distribution"
            chart.height = 10
            chart.width = 15
            data_ref = Reference(ws, min_col=2, min_row=current_row-len(income_data)-1, max_row=current_row-2)
            cats_ref = Reference(ws, min_col=1, min_row=current_row-len(income_data), max_row=current_row-2)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, f'F{current_row-len(income_data)-1}')
        
        current_row += 2
        
        # Additional Economic Indicators
        ws.merge_cells(f'A{current_row}:D{current_row}')
        additional_title = ws.cell(row=current_row, column=1)
        additional_title.value = "Economic Indicators"
        additional_title.font = self.subheader_font
        
        current_row += 1
        
        # Calculate additional indicators
        base_qs = self.data['demographics']  # Using demographics as proxy
        total_applicants = self.data['summary']['total_applicants']
        
        # Sample additional indicators
        indicators_data = [
            ["Below Poverty Threshold (<10k)", 
             econ['income_distribution'][0]['count'] if econ['income_distribution'] else 0],
            ["Middle Income (20k-50k)", 
             sum(item['count'] for item in econ['income_distribution'][2:5]) if len(econ['income_distribution']) >= 5 else 0],
            ["Above Average (>50k)", 
             sum(item['count'] for item in econ['income_distribution'][5:]) if len(econ['income_distribution']) > 5 else 0],
        ]
        
        current_row = self._create_data_table(ws, current_row, 
                                              ['Economic Indicator', 'Count'], 
                                              indicators_data, [30, 15])


    def _create_trends_sheet(self):
        """Sheet 4: Trends Analysis"""
        ws = self.wb.create_sheet("4. Trends")
        
        current_row = self._add_report_header(ws, "TRENDS ANALYSIS")
        
        trends = self.data['trends']
        
        # Monthly Trends
        ws.merge_cells(f'A{current_row}:D{current_row}')
        monthly_title = ws.cell(row=current_row, column=1)
        monthly_title.value = "Monthly Application Trends (Last 12 Months)"
        monthly_title.font = self.subheader_font
        
        current_row += 1
        
        if trends['monthly']:
            monthly_data = [[item['month'].strftime('%b %Y') if item['month'] else 'N/A', 
                           item['count']] for item in trends['monthly']]
            
            table_start = current_row
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Month', 'Applications'], 
                                                  monthly_data, [20, 15])
            
            # Line chart
            line_chart = LineChart()
            line_chart.title = "Monthly Trend"
            line_chart.style = 12
            line_chart.height = 10
            line_chart.width = 20
            
            data_ref = Reference(ws, min_col=2, min_row=table_start, max_row=table_start+len(monthly_data))
            cats_ref = Reference(ws, min_col=1, min_row=table_start+1, max_row=table_start+len(monthly_data))
            line_chart.add_data(data_ref, titles_from_data=True)
            line_chart.set_categories(cats_ref)
            
            ws.add_chart(line_chart, f'F{table_start}')
        
        current_row += 2
        
        # Yearly Trends
        ws.merge_cells(f'A{current_row}:D{current_row}')
        yearly_title = ws.cell(row=current_row, column=1)
        yearly_title.value = "Yearly Application Trends"
        yearly_title.font = self.subheader_font
        
        current_row += 1
        
        if trends['yearly']:
            yearly_data = [[item['year'], item['count']] for item in trends['yearly']]
            
            table_start = current_row
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Year', 'Applications'], 
                                                  yearly_data, [20, 15])
            
            # Bar chart
            bar_chart = BarChart()
            bar_chart.title = "Yearly Comparison"
            bar_chart.height = 8
            bar_chart.width = 15
            
            data_ref = Reference(ws, min_col=2, min_row=table_start, max_row=table_start+len(yearly_data))
            cats_ref = Reference(ws, min_col=1, min_row=table_start+1, max_row=table_start+len(yearly_data))
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            
            ws.add_chart(bar_chart, f'F{table_start}')
        
        current_row += 2
        
        # Assistance Type Distribution
        ws.merge_cells(f'A{current_row}:D{current_row}')
        assist_title = ws.cell(row=current_row, column=1)
        assist_title.value = "Applications by Assistance Type"
        assist_title.font = self.subheader_font
        
        current_row += 1
        
        if trends['by_assistance']:
            assist_data = [[item['type_of_assistance'], item['count']] 
                          for item in trends['by_assistance']]
            
            table_start = current_row
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Assistance Type', 'Applications'], 
                                                  assist_data, [35, 15])
            
            # Combo chart (bar + line)
            bar_chart = BarChart()
            bar_chart.title = "Assistance Type Distribution"
            bar_chart.height = 12
            bar_chart.width = 20
            
            data_ref = Reference(ws, min_col=2, min_row=table_start, max_row=table_start+len(assist_data))
            cats_ref = Reference(ws, min_col=1, min_row=table_start+1, max_row=table_start+len(assist_data))
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            
            ws.add_chart(bar_chart, f'F{table_start}')
        
        current_row += 2
        
        # Assistance Types Over Time (if data available)
        if trends.get('assistance_over_time'):
            ws.merge_cells(f'A{current_row}:D{current_row}')
            overtime_title = ws.cell(row=current_row, column=1)
            overtime_title.value = "Assistance Types Trend Over Time"
            overtime_title.font = self.subheader_font
            
            current_row += 1

    def _create_performance_sheet(self):
        """Sheet 5: Performance Metrics"""
        ws = self.wb.create_sheet("5. Performance")
        
        current_row = self._add_report_header(ws, "PERFORMANCE METRICS")
        
        perf = self.data['performance']
        
        # Staff Productivity
        ws.merge_cells(f'A{current_row}:D{current_row}')
        staff_title = ws.cell(row=current_row, column=1)
        staff_title.value = "Staff Productivity (Top 10)"
        staff_title.font = self.subheader_font
        
        current_row += 1
        
        if perf['staff_productivity']:
            staff_data = [[item['staff__username'], item['count']] 
                         for item in perf['staff_productivity'][:10]]
            
            table_start = current_row
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Staff Member', 'Applications Processed'], 
                                                  staff_data, [25, 20])
            
            # Highlight top and bottom performers
            # Top performer (green)
            top_row = table_start + 1
            for col in range(1, 3):
                ws.cell(row=top_row, column=col).fill = PatternFill(
                    start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'
                )
                ws.cell(row=top_row, column=col).font = Font(bold=True, color='2E7D32')
            
            # Bottom performer (light red) if more than 3 staff
            if len(staff_data) > 3:
                bottom_row = table_start + len(staff_data)
                for col in range(1, 3):
                    ws.cell(row=bottom_row, column=col).fill = PatternFill(
                        start_color='FFCCBC', end_color='FFCCBC', fill_type='solid'
                    )
            
            # Bar chart
            chart = BarChart()
            chart.title = "Staff Productivity Comparison"
            chart.height = 12
            chart.width = 18
            
            data_ref = Reference(ws, min_col=2, min_row=table_start, max_row=table_start+len(staff_data))
            cats_ref = Reference(ws, min_col=1, min_row=table_start+1, max_row=table_start+len(staff_data))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            ws.add_chart(chart, f'F{table_start}')
        
        current_row += 2
        
        # Processing Time by Assistance Type
        ws.merge_cells(f'A{current_row}:D{current_row}')
        process_title = ws.cell(row=current_row, column=1)
        process_title.value = "Average Processing Time by Assistance Type"
        process_title.font = self.subheader_font
        
        current_row += 1
        
        if perf['processing_by_type']:
            process_data = [[item['type_of_assistance'], 
                           f"{item.get('avg_minutes', 0):.1f}"] 
                          for item in perf['processing_by_type']]
            
            table_start = current_row
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Assistance Type', 'Avg Time (minutes)'], 
                                                  process_data, [35, 20])
            
            # Highlight fastest (green) and slowest (red)
            if len(process_data) > 1:
                # Find fastest and slowest
                times = [float(item[1]) for item in process_data]
                fastest_idx = times.index(min(times))
                slowest_idx = times.index(max(times))
                
                # Fastest (green)
                fastest_row = table_start + 1 + fastest_idx
                for col in range(1, 3):
                    ws.cell(row=fastest_row, column=col).fill = PatternFill(
                        start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'
                    )
                    ws.cell(row=fastest_row, column=col).font = Font(bold=True, color='2E7D32')
                
                # Slowest (red)
                slowest_row = table_start + 1 + slowest_idx
                for col in range(1, 3):
                    ws.cell(row=slowest_row, column=col).fill = PatternFill(
                        start_color='FFCCBC', end_color='FFCCBC', fill_type='solid'
                    )
                    ws.cell(row=slowest_row, column=col).font = Font(bold=True, color='C62828')
            
            # Bar chart
            chart = BarChart()
            chart.title = "Processing Time Comparison"
            chart.height = 10
            chart.width = 18
            
            data_ref = Reference(ws, min_col=2, min_row=table_start, max_row=table_start+len(process_data))
            cats_ref = Reference(ws, min_col=1, min_row=table_start+1, max_row=table_start+len(process_data))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            ws.add_chart(chart, f'F{table_start}')
        
        current_row += 2
        
        # Activity Heatmap
        ws.merge_cells(f'A{current_row}:D{current_row}')
        heat_title = ws.cell(row=current_row, column=1)
        heat_title.value = "Application Activity by Hour of Day"
        heat_title.font = self.subheader_font
        
        current_row += 1
        
        if perf['activity_heatmap']:
            heat_data = [[f"{item['hour']:02d}:00", item['count']] 
                        for item in perf['activity_heatmap']]
            
            table_start = current_row
            current_row = self._create_data_table(ws, current_row, 
                                                  ['Hour of Day', 'Applications'], 
                                                  heat_data, [20, 15])
            
            # Apply heatmap color scale
            max_count = max(item['count'] for item in perf['activity_heatmap'])
            for idx, item in enumerate(perf['activity_heatmap'], start=1):
                intensity = item['count'] / max_count if max_count > 0 else 0
                color_val = int(255 - (intensity * 200))  # Darker blue for higher values
                color_hex = f"{color_val:02x}{color_val:02x}ff"
                
                ws.cell(row=table_start+idx, column=2).fill = PatternFill(
                    start_color=color_hex, end_color=color_hex, fill_type='solid'
                )
                if intensity > 0.5:
                    ws.cell(row=table_start+idx, column=2).font = Font(color='FFFFFF', bold=True)
            
            # Line chart
            line_chart = LineChart()
            line_chart.title = "Hourly Activity Pattern"
            line_chart.height = 10
            line_chart.width = 18
            
            data_ref = Reference(ws, min_col=2, min_row=table_start, max_row=table_start+len(heat_data))
            cats_ref = Reference(ws, min_col=1, min_row=table_start+1, max_row=table_start+len(heat_data))
            line_chart.add_data(data_ref, titles_from_data=True)
            line_chart.set_categories(cats_ref)
            
            ws.add_chart(line_chart, f'F{table_start}')
        
        current_row += 2
        
        # Performance Summary KPIs
        ws.merge_cells(f'A{current_row}:H{current_row}')
        kpi_header = ws.cell(row=current_row, column=1)
        kpi_header.value = "Performance Summary"
        kpi_header.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        kpi_header.fill = self.accent_fill
        kpi_header.alignment = self.center_align
        
        current_row += 2
        
        # Calculate performance KPIs
        avg_processing = self.data['summary']['avg_processing_minutes']
        total_staff = len(perf['staff_productivity']) if perf['staff_productivity'] else 0
        total_processed = sum(item['count'] for item in perf['staff_productivity']) if perf['staff_productivity'] else 0
        avg_per_staff = total_processed / total_staff if total_staff > 0 else 0
        
        peak_hour = max(perf['activity_heatmap'], key=lambda x: x['count'])['hour'] if perf['activity_heatmap'] else 0
        
        kpi_summary = [
            ["Metric", "Value", "Status"],
            ["Average Processing Time", f"{avg_processing:.1f} min", 
             "✓ Good" if avg_processing < 10 else "⚠ Needs Improvement"],
            ["Total Staff Active", total_staff, ""],
            ["Avg Applications per Staff", f"{avg_per_staff:.1f}", ""],
            ["Peak Activity Hour", f"{peak_hour:02d}:00", ""],
        ]
        
        current_row = self._create_data_table(ws, current_row, 
                                              kpi_summary[0], 
                                              kpi_summary[1:], [30, 20, 25])

    def _create_insights_recommendations_sheet(self):
        """Sheet 6: Insights and Recommendations"""
        ws = self.wb.create_sheet("6. Insights & Recommendations")
        
        current_row = self._add_report_header(ws, "INSIGHTS & RECOMMENDATIONS")
        
        insights = self.insights
        
        # Geographic Insights
        ws.merge_cells(f'A{current_row}:H{current_row}')
        geo_header = ws.cell(row=current_row, column=1)
        geo_header.value = "GEOGRAPHIC INSIGHTS"
        geo_header.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        geo_header.fill = self.header_fill
        geo_header.alignment = self.center_align
        
        current_row += 2
        
        if insights.get('geographic'):
            for insight in insights['geographic']:
                ws.merge_cells(f'A{current_row}:H{current_row}')
                cell = ws.cell(row=current_row, column=1)
                cell.value = insight
                cell.font = self.body_font
                cell.alignment = self.left_align
                cell.border = self.thin_border
                
                # Add bullet styling
                if insight.strip().startswith('•'):
                    cell.fill = self.light_fill
                
                ws.row_dimensions[current_row].height = 25
                current_row += 1
        
        current_row += 1
        
        # Demographic Insights
        ws.merge_cells(f'A{current_row}:H{current_row}')
        demo_header = ws.cell(row=current_row, column=1)
        demo_header.value = "DEMOGRAPHIC INSIGHTS"
        demo_header.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        demo_header.fill = self.header_fill
        demo_header.alignment = self.center_align
        
        current_row += 2
        
        if insights.get('demographic'):
            for insight in insights['demographic']:
                ws.merge_cells(f'A{current_row}:H{current_row}')
                cell = ws.cell(row=current_row, column=1)
                cell.value = insight
                cell.font = self.body_font
                cell.alignment = self.left_align
                cell.border = self.thin_border
                
                if insight.strip().startswith('•'):
                    cell.fill = self.light_fill
                
                ws.row_dimensions[current_row].height = 25
                current_row += 1
        
        current_row += 1
        
        # Trend Insights
        ws.merge_cells(f'A{current_row}:H{current_row}')
        trend_header = ws.cell(row=current_row, column=1)
        trend_header.value = "TREND INSIGHTS"
        trend_header.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        trend_header.fill = self.header_fill
        trend_header.alignment = self.center_align
        
        current_row += 2
        
        if insights.get('trends'):
            for insight in insights['trends']:
                ws.merge_cells(f'A{current_row}:H{current_row}')
                cell = ws.cell(row=current_row, column=1)
                cell.value = insight
                cell.font = self.body_font
                cell.alignment = self.left_align
                cell.border = self.thin_border
                
                if insight.strip().startswith('•'):
                    cell.fill = self.light_fill
                
                ws.row_dimensions[current_row].height = 25
                current_row += 1
        
        current_row += 1
        
        # Performance Insights
        ws.merge_cells(f'A{current_row}:H{current_row}')
        perf_header = ws.cell(row=current_row, column=1)
        perf_header.value = "PERFORMANCE INSIGHTS"
        perf_header.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        perf_header.fill = self.header_fill
        perf_header.alignment = self.center_align
        
        current_row += 2
        
        if insights.get('performance'):
            for insight in insights['performance']:
                ws.merge_cells(f'A{current_row}:H{current_row}')
                cell = ws.cell(row=current_row, column=1)
                cell.value = insight
                cell.font = self.body_font
                cell.alignment = self.left_align
                cell.border = self.thin_border
                
                if insight.strip().startswith('•'):
                    cell.fill = self.light_fill
                
                ws.row_dimensions[current_row].height = 25
                current_row += 1
        
        current_row += 2
        
        # Recommendations Section (Priority-based)
        ws.merge_cells(f'A{current_row}:H{current_row}')
        rec_header = ws.cell(row=current_row, column=1)
        rec_header.value = "ACTIONABLE RECOMMENDATIONS"
        rec_header.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        rec_header.fill = PatternFill(start_color='f44336', end_color='f44336', fill_type='solid')
        rec_header.alignment = self.center_align
        
        current_row += 2
        
        # Add priority levels to recommendations
        if insights.get('recommendations'):
            # Create table with priority levels
            rec_table_headers = ["Priority", "Recommendation", "Impact"]
            
            ws.cell(row=current_row, column=1, value=rec_table_headers[0]).font = self.header_font
            ws.cell(row=current_row, column=1).fill = self.header_fill
            ws.cell(row=current_row, column=1).alignment = self.center_align
            ws.cell(row=current_row, column=1).border = self.thin_border
            
            ws.merge_cells(f'B{current_row}:G{current_row}')
            ws.cell(row=current_row, column=2, value=rec_table_headers[1]).font = self.header_font
            ws.cell(row=current_row, column=2).fill = self.header_fill
            ws.cell(row=current_row, column=2).alignment = self.center_align
            ws.cell(row=current_row, column=2).border = self.thin_border
            
            ws.cell(row=current_row, column=8, value=rec_table_headers[2]).font = self.header_font
            ws.cell(row=current_row, column=8).fill = self.header_fill
            ws.cell(row=current_row, column=8).alignment = self.center_align
            ws.cell(row=current_row, column=8).border = self.thin_border
            
            current_row += 1
            
            # Assign priorities (you can customize logic)
            priorities = ["HIGH", "HIGH", "MEDIUM"]
            impacts = ["High", "High", "Medium"]
            
            for idx, rec in enumerate(insights['recommendations']):
                priority = priorities[idx] if idx < len(priorities) else "LOW"
                impact = impacts[idx] if idx < len(impacts) else "Low"
                
                # Priority cell
                priority_cell = ws.cell(row=current_row, column=1, value=priority)
                priority_cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                priority_cell.alignment = self.center_align
                priority_cell.border = self.thin_border
                
                if priority == "HIGH":
                    priority_cell.fill = PatternFill(start_color='f44336', end_color='f44336', fill_type='solid')
                elif priority == "MEDIUM":
                    priority_cell.fill = PatternFill(start_color='ff9800', end_color='ff9800', fill_type='solid')
                else:
                    priority_cell.fill = PatternFill(start_color='4caf50', end_color='4caf50', fill_type='solid')
                
                # Recommendation cell
                ws.merge_cells(f'B{current_row}:G{current_row}')
                rec_cell = ws.cell(row=current_row, column=2, value=rec)
                rec_cell.font = self.body_font
                rec_cell.alignment = self.left_align
                rec_cell.border = self.thin_border
                ws.row_dimensions[current_row].height = 35
                
                # Impact cell
                impact_cell = ws.cell(row=current_row, column=8, value=impact)
                impact_cell.font = self.body_font
                impact_cell.alignment = self.center_align
                impact_cell.border = self.thin_border
                
                current_row += 1
        
        current_row += 2
        
        # Data Sources & Methodology
        ws.merge_cells(f'A{current_row}:H{current_row}')
        method_header = ws.cell(row=current_row, column=1)
        method_header.value = "DATA SOURCES & METHODOLOGY"
        method_header.font = Font(name='Calibri', size=11, bold=True, color='666666')
        method_header.fill = self.gray_fill
        method_header.alignment = self.center_align
        
        current_row += 2
        
        methodology_text = f"""
Data Source: DSWD Applicant Management System
Analysis Period: {self.filters.get('start_date', 'All Records')} to {self.filters.get('end_date', 'Present')}
Total Records Analyzed: {self.data['summary']['total_applicants']:,} applicants
Methodology: Statistical analysis using Python analytics engine with Django ORM queries
Filters Applied: {'Cities: ' + ', '.join(self.filters.get('cities', [])) if self.filters.get('cities') else 'All Cities'} | {'Assistance Types: ' + ', '.join(self.filters.get('assistance_types', [])) if self.filters.get('assistance_types') else 'All Types'}
        """.strip()
        
        ws.merge_cells(f'A{current_row}:H{current_row+4}')
        method_cell = ws.cell(row=current_row, column=1)
        method_cell.value = methodology_text
        method_cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
        method_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        method_cell.border = self.thin_border
        ws.row_dimensions[current_row].height = 80
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['H'].width = 12

    def generate(self, output_path):
        """Generate the complete 6-sheet Excel report"""
        try:
            # Create all 6 sheets in order
            self._create_summary_sheet()
            self._create_geographic_sheet()
            self._create_demographics_economics_sheet()
            self._create_trends_sheet()
            self._create_performance_sheet()
            self._create_insights_recommendations_sheet()
            
            # Save workbook
            self.wb.save(output_path)
            
            return True
            
        except Exception as e:
            print(f"Error generating Excel report: {str(e)}")
            import traceback
            traceback.print_exc()
            return False



class ExportOrchestrator:
    """Orchestrates the entire export process"""


    def get_cached_result(self, filters: dict):
        """
        Caches analytics results for 5 minutes to avoid recomputing
        """
        key = "analytics_export_" + hashlib.md5(json.dumps(filters, sort_keys=True).encode()).hexdigest()

        cached = cache.get(key)
        if cached:
            return cached

        # Compute normally
        result = self.compute_all_analytics()  # whatever method produces all metrics
        cache.set(key, result, timeout=300)  # 5 minutes

        return result

    
    def __init__(self, filters, user, format_type='both'):
        self.filters = filters
        self.user = user
        self.format_type = format_type
        self.branding = filters.get('branding', {})
        
        # Set default branding
        if 'organization_name' not in self.branding:
            self.branding['organization_name'] = 'DSWD Quezon Province'
        if 'primary_color' not in self.branding:
            self.branding['primary_color'] = '#0066cc'
    
    def generate_report(self):
        """Generate the complete analytics report and return as base64"""
        try:
            # Step 1: Collect all data
            collector = AnalyticsDataCollector(self.filters)
            data = collector.collect_all()
            
            # Step 2: Generate insights
            insights_gen = InsightsGenerator(data)
            insights = insights_gen.generate_all()
            
            # Step 3: Generate charts
            chart_gen = ChartGenerator()
            charts = chart_gen.generate_all_charts(data)
            
            # Step 4: Prepare results
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            results = {
                'success': True,
                'files': {},
                'metadata': {
                    'generated_at': datetime.now().isoformat(),
                    'generated_by': self.user.username,
                    'total_applicants': data['summary']['total_applicants'],
                    'date_range': f"{self.filters.get('start_date', 'All')} to {self.filters.get('end_date', 'All')}"
                }
            }
            
            # Step 5: Generate PDF in memory
            if self.format_type in ['pdf', 'both']:
                pdf_buffer = BytesIO()
                pdf_gen = PDFReportGenerator(data, insights, charts, self.branding, self.filters)
                
                # Create document
                doc = SimpleDocTemplate(
                    pdf_buffer,
                    pagesize=A4,
                    rightMargin=0.75*inch,
                    leftMargin=0.75*inch,
                    topMargin=0.9*inch,
                    bottomMargin=0.8*inch
                )
                
                story = []
                
                # Build all sections
                story.extend(pdf_gen._create_cover_page())
                story.extend(pdf_gen._create_table_of_contents())
                story.extend(pdf_gen._create_executive_summary())
                story.extend(pdf_gen._create_geographic_analysis())
                story.extend(pdf_gen._create_demographics_economics())
                story.extend(pdf_gen._create_trends_analysis())
                story.extend(pdf_gen._create_performance_metrics())
                story.extend(pdf_gen._create_insights_recommendations())
                story.extend(pdf_gen._create_figures_appendix())
                story.extend(pdf_gen._create_tables_appendix())
                story.extend(pdf_gen._create_methodology_appendix())
                story.extend(pdf_gen._create_charts_section())
                
                # Build PDF
                doc.build(story, 
                        onFirstPage=pdf_gen._corporate_header_footer, 
                        onLaterPages=pdf_gen._corporate_header_footer)
                
                # Convert to base64
                pdf_buffer.seek(0)
                pdf_base64 = base64.b64encode(pdf_buffer.read()).decode('utf-8')
                
                results['files']['pdf'] = {
                    'filename': f'analytics_report_{timestamp}.pdf',
                    'base64': pdf_base64,
                    'content_type': 'application/pdf'
                }
                
                pdf_buffer.close()
            
            # Step 6: Generate Excel in memory
            if self.format_type in ['excel', 'both']:
                excel_buffer = BytesIO()
                excel_gen = ExcelReportGenerator(data, insights, charts, self.branding, self.filters)
                
                # Generate all sheets
                excel_gen.generate(excel_buffer)
                
                # Convert to base64
                excel_buffer.seek(0)
                excel_base64 = base64.b64encode(excel_buffer.read()).decode('utf-8')
                
                results['files']['excel'] = {
                    'filename': f'analytics_report_{timestamp}.xlsx',
                    'base64': excel_base64,
                    'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
                
                excel_buffer.close()

            
            # Close all matplotlib figures
            plt.close('all')
            
            return results
            
        except Exception as e:
            import traceback
            return {
                'success': False,
                'error': str(e),
                'traceback': traceback.format_exc()
            }